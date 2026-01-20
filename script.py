import os
import sys
import re
import math
import subprocess
import json
import time
import shutil

import pandas as pd
from typing import List, Set, Tuple, Optional, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font


# ===================== FIX DE FONTE =====================
# Definições globais de fonte (antes de qualquer style.configure)

import tkinter as tk
from tkinter import ttk

fonte_padrao = ("Segoe UI", 10)
fonte_label = ("Segoe UI", 10)
fonte_titulo = ("Segoe UI Semibold", 12)
fonte_botao = ("Segoe UI Semibold", 10)
# ===== DEFINIÇÕES COMPLETAS DE FONTES (FIX FINAL) =====
fonte_entry = ("Segoe UI", 10)
fonte_combo = ("Segoe UI", 10)
fonte_checkbox = ("Segoe UI", 10)
# =====================================================


# ========================================================
# O restante do código permanece igual, apenas garantindo
# que fonte_label exista antes de ser usada no ttk.Style
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from sqlalchemy import create_engine, text as sql_text
from sqlalchemy.engine import URL

# (NOVO) gráficos embutidos no Tkinter
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# -------------------- CORES E ESTILO ------------------------
BG_PRINCIPAL = "#111827"   # fundo geral
BG_FRAME = "#1F2933"       # fundo dos blocos
FG_TEXTO = "#E5E7EB"       # texto principal
FG_SECUNDARIO = "#9CA3AF"  # texto secundário
ACCENT_GREEN = "#22C55E"   # botão de ação
ACCENT_BLUE = "#3B82F6"    # botão secundário
ACCENT_ORANGE = "#F97316"  # botão de pasta
BORDER_COR = "#374151"     # bordas / contornos
INPUT_BG = "#020617"       # campos de texto
INPUT_FG = "#F9FAFB"       # texto dos campos

# -------------------- CONSTANTES LIMPEZA --------------------
PHONE_MIN_LEN = 8

# (NOVO) DDD por Estado (map invertido para lookup rápido)
DDD_ESTADOS: Dict[str, List[str]] = {
    "AC": ["68"],
    "AL": ["82"],
    "AP": ["96"],
    "AM": ["92", "97"],
    "BA": ["71", "73", "74", "75", "77"],
    "CE": ["85", "88"],
    "DF": ["61"],
    "ES": ["27", "28"],
    "GO": ["62", "64"],
    "MA": ["98", "99"],
    "MT": ["65", "66"],
    "MS": ["67"],
    "MG": ["31", "32", "33", "34", "35", "37", "38"],
    "PA": ["91", "93", "94"],
    "PB": ["83"],
    "PR": ["41", "42", "43", "44", "45", "46"],
    "PE": ["81", "87"],
    "PI": ["86", "89"],
    "RJ": ["21", "22", "24"],
    "RN": ["84"],
    "RS": ["51", "53", "54", "55"],
    "RO": ["69"],
    "RR": ["95"],
    "SC": ["47", "48", "49"],
    "SP": ["11", "12", "13", "14", "15", "16", "17", "18", "19"],
    "SE": ["79"],
    "TO": ["63"],
}
DDD_TO_UF: Dict[str, str] = {}
for uf, ddds in DDD_ESTADOS.items():
    for ddd in ddds:
        DDD_TO_UF[ddd] = uf

# -------------------- CONFIG BANCO -------------------------
DB_CONFIG_FILE = "db_config.json"
db_engine = None
db_connected = False


# =======================================================================
#           FUNÇÕES COMPARTILHADAS / UTILITÁRIOS (LÓGICA)
# =======================================================================

def normalize_col_name(name: str) -> str:
    return re.sub(r"[^0-9a-zA-Z]+", "", str(name)).strip().lower()

def read_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in ['.xls', '.xlsx']:
        return pd.read_excel(path, dtype=str)
    elif ext in ['.csv', '.txt']:
        return pd.read_csv(path, dtype=str, sep=None, engine='python')
    else:
        raise ValueError('Formato não suportado: ' + ext)

def save_to_excel(df: pd.DataFrame, path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dados'
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(path)

def normalize_cnpj(c):
    digits = re.sub(r'\D', '', str(c or ''))
    if len(digits) > 14:
        digits = digits[-14:]
    return digits.zfill(14) if digits else None

def split_telefones_field(field: str) -> Tuple[Optional[str], Optional[str]]:
    if pd.isna(field):
        return (None, None)
    s = str(field)
    parts = re.split(r'[;,/\|\s]+', s)
    phones = [re.sub(r'\D', '', p) for p in parts if re.sub(r'\D', '', p)]
    return (phones + [None, None])[:2]

def normalize_phone(num: str, *, strip55: bool = False, add9: bool = False, add55: bool = False) -> str:
    """
    Normaliza telefone para somente dígitos e aplica regras opcionais:

    - strip55=True: se começar com '55' (código país), remove.
    - add9=True: se tiver 10 dígitos (DD + 8), insere '9' após o DDD -> 11 dígitos.
      (Ex.: 1999659233 -> 19999659233)
    - add55=True: se NÃO começar com '55', adiciona '55' no início.
    """
    digits = re.sub(r'\D', '', str(num or ''))

    if strip55 and digits.startswith('55') and len(digits) >= 12:
        digits = digits[2:]

    # adiciona 9 após DDD quando formato for DD + 8
    if add9 and len(digits) == 10:
        digits = digits[:2] + '9' + digits[2:]

    if add55 and digits and not digits.startswith('55'):
        digits = '55' + digits

    return digits


def is_invalid_phone(num: str) -> bool:
    if not num:
        return True
    digits = re.sub(r'\D', '', str(num))
    if len(digits) < PHONE_MIN_LEN:
        return True
    if len(set(digits)) == 1:
        return True
    return False

def clean_razao_social(s: str) -> str:
    """
    Remove da Razao Social:
    0-9 - = ' " , . ; [ ] : { } ! @ # $ % & ( ) _ +
    Mantém letras e espaços.
    """
    if pd.isna(s):
        return ''
    pattern = r"[0-9\-\=\'\",.;\[\]:\{\}!@#$%&\(\)_\+]"
    return re.sub(pattern, "", str(s))

def pick_col(normals: dict, candidates: list):
    for cand in candidates:
        key = normalize_col_name(cand)
        if key in normals:
            return normals[key]
        for norm, orig in normals.items():
            if key in norm:
                return orig
    return None

def safe_remove_file(path: str):
    try:
        if os.path.isfile(path):
            os.remove(path)
    except:
        pass

def _extract_ddd_from_phone(phone: str) -> Optional[str]:
    """
    Regra: DDD são os 2 primeiros dígitos do telefone (conforme seu pedido).
    """
    p = normalize_phone(phone)
    if p.startswith('55') and len(p) >= 4:
        p = p[2:]
    if not p or len(p) < 2:
        return None
    return p[:2]

def uf_from_phone(tel1: str, tel2: str) -> str:
    """
    Define UF usando primeiro DDD disponível (Telefone1, senão Telefone2).
    Se não achar, retorna "??".
    """
    ddd = _extract_ddd_from_phone(tel1) or _extract_ddd_from_phone(tel2)
    if not ddd:
        return "??"
    return DDD_TO_UF.get(ddd, "??")

def _append_reason(reason_str: str, reason: str) -> str:
    reason = (reason or "").strip()
    if not reason:
        return reason_str or ""
    if not reason_str:
        return reason
    # evita duplicar
    parts = [p.strip() for p in reason_str.split(" | ") if p.strip()]
    if reason not in parts:
        parts.append(reason)
    return " | ".join(parts)


# =======================================================================
#           FUNÇÕES DA INTERFACE PROCV B2B
# =======================================================================

def abrir_pasta():
    if caminho_arquivo_saida.get() != "":
        pasta = os.path.dirname(caminho_arquivo_saida.get())
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", pasta])
            elif os.name == "nt":
                subprocess.Popen(f'explorer "{pasta}"')
            else:
                subprocess.Popen(["xdg-open", pasta])
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a pasta.\n\n{e}")
    else:
        messagebox.showwarning("Aviso", "Nenhum arquivo foi gerado ainda.")

def selecionar_pasta_saida():
    pasta = filedialog.askdirectory(title="Selecione a pasta onde salvar o arquivo")
    if pasta:
        pasta_saida.set(pasta)
    else:
        messagebox.showinfo("Aviso", "Nenhuma pasta selecionada.")

def carregar_colunas():
    try:
        arquivo = caminho_arquivo.get()
        if not arquivo:
            messagebox.showwarning("Aviso", "Selecione um arquivo primeiro.")
            return

        if arquivo.endswith(".xlsx"):
            df = pd.read_excel(arquivo)
        elif arquivo.endswith(".csv"):
            df = pd.read_csv(arquivo, sep=";", encoding="utf-8")
        else:
            messagebox.showerror("Erro", "Selecione um arquivo CSV ou XLSX.")
            return

        colunas = list(df.columns)
        combo_colA["values"] = colunas
        combo_colB["values"] = colunas
        messagebox.showinfo("OK", "Colunas carregadas com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível carregar colunas.\n\n{e}")

def executar_comparacao():
    try:
        progress["value"] = 0
        janela.update_idletasks()

        opcao = combo_opcao.get()
        if opcao == "":
            messagebox.showwarning("Aviso", "Selecione o tipo de comparação.")
            return

        if caminho_arquivo.get() == "":
            messagebox.showwarning("Aviso", "Selecione um arquivo.")
            return

        if pasta_saida.get() == "":
            messagebox.showwarning("Aviso", "Selecione onde salvar o arquivo final.")
            return

        arquivo = caminho_arquivo.get()
        pasta_destino = pasta_saida.get()
        ext = arquivo.split(".")[-1].lower()
        progress["value"] = 10
        janela.update_idletasks()

        if ext == "xlsx":
            df = pd.read_excel(arquivo)
        elif ext == "csv":
            df = pd.read_csv(arquivo, sep=";", encoding="utf-8")
        else:
            messagebox.showerror("Erro", "Formato não suportado. Use CSV ou XLSX.")
            return

        progress["value"] = 30
        janela.update_idletasks()

        colA = combo_colA.get()
        colB = combo_colB.get()
        if colA == "" or colB == "":
            messagebox.showerror("Erro", "Selecione as colunas para comparação.")
            return

        if opcao == "O que tem na A e não tem na B":
            resultado = df[~df[colA].isin(df[colB])][colA]
            tipo = f"{colA}NAO_ESTA_EM{colB}"
            coluna_base = colA
            outra_coluna = colB
        else:
            resultado = df[~df[colB].isin(df[colA])][colB]
            tipo = f"{colB}NAO_ESTA_EM{colA}"
            coluna_base = colB
            outra_coluna = colA

        progress["value"] = 50
        janela.update_idletasks()

        df[tipo] = ""
        df.loc[~df[coluna_base].isin(df[outra_coluna]), tipo] = df[coluna_base]

        nome_arquivo_saida = f"resultado_{tipo}.xlsx"
        arquivo_saida = os.path.join(pasta_destino, nome_arquivo_saida)
        df.to_excel(arquivo_saida, index=False)
        caminho_arquivo_saida.set(arquivo_saida)
        progress["value"] = 70
        janela.update_idletasks()

        wb = load_workbook(arquivo_saida)
        ws = wb.active
        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        col_base_idx = df.columns.get_loc(coluna_base) + 1
        col_res_idx = df.columns.get_loc(tipo) + 1

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_base_idx)
            cell_res = ws.cell(row=row, column=col_res_idx)
            if cell_res.value not in ("", None):
                cell.fill = fill_amarelo
                cell_res.fill = fill_amarelo

        wb.save(arquivo_saida)
        progress["value"] = 90
        janela.update_idletasks()

        relatorio = f"""
PROCESSO COMPLETO

Arquivo analisado: {arquivo}
Arquivo gerado em: {arquivo_saida}

Tipo de comparação: {tipo}

Linhas analisadas: {len(df)}
Itens encontrados: {len(resultado)}

Lista dos itens encontrados:
{resultado.to_list()}
"""
        txt_relatorio.delete("1.0", tk.END)
        txt_relatorio.insert(tk.END, relatorio)
        progress["value"] = 100
        janela.update_idletasks()
        messagebox.showinfo("Concluído", "Comparação finalizada com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", str(e))


# =======================================================================
#           FUNÇÕES DA LIMPEZA DE DADOS (NOVA REGRA)
# =======================================================================

def log_limpeza(msg: str):
    txt_log_limpeza.insert(tk.END, msg + "\n")
    txt_log_limpeza.see(tk.END)
    janela.update_idletasks()

def abrir_pasta_limpeza():
    pasta = out_dir_limpeza.get()
    if not pasta:
        messagebox.showwarning("Aviso", "Nenhuma pasta de saída selecionada.")
        return
    try:
        if sys.platform == "darwin":
            subprocess.Popen(["open", pasta])
        elif os.name == "nt":
            subprocess.Popen(f'explorer "{pasta}"')
        else:
            subprocess.Popen(["xdg-open", pasta])
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir a pasta.\n\n{e}")

def _load_phone_set(path: str, label: str, *, strip55: bool = False) -> Set[str]:
    path = (path or "").strip()
    if not path:
        return set()
    if not os.path.isfile(path):
        log_limpeza(f"⚠️ {label}: arquivo não encontrado: {path}")
        return set()

    try:
        tdf = read_table(path)
        tel_col = next((c for c in tdf.columns if "tel" in normalize_col_name(c)), None)
        if not tel_col:
            tel_col = tdf.columns[0]
        s = set(tdf[tel_col].astype(str).map(lambda x: normalize_phone(x, strip55=strip55)))
        s = {x for x in s if x}
        log_limpeza(f"✅ {label}: {len(s)} telefones carregados ({os.path.basename(path)})")
        return s
    except Exception as e:
        log_limpeza(f"❌ {label}: erro ao ler: {e}")
        return set()

def escanear_colunas_limpeza():
    try:
        in_path = base_empresas_path.get().strip()
        if not in_path:
            messagebox.showwarning("Aviso", 'Selecione o arquivo "empresas bruto" primeiro.')
            return

        log_limpeza("🔎 Escaneando colunas do arquivo...")
        df = read_table(in_path)
        cols = list(df.columns)
        if not cols:
            messagebox.showerror("Erro", "Não foi possível identificar colunas no arquivo.")
            return

        combo_limpeza_razao["values"] = cols
        combo_limpeza_tel["values"] = cols
        combo_limpeza_email["values"] = cols
        combo_limpeza_cnpj["values"] = cols

        normals = {normalize_col_name(c): c for c in cols}

        def suggest(keys):
            for k in keys:
                kk = normalize_col_name(k)
                if kk in normals:
                    return normals[kk]
            for c in cols:
                cn = normalize_col_name(c)
                if any(normalize_col_name(k) in cn for k in keys):
                    return c
            return ""

        limpeza_col_razao.set(suggest(["razao social", "razao", "nome empresa", "empresa"]))
        limpeza_col_tel.set(suggest(["telefones", "telefone", "tel", "fone", "celular"]))
        limpeza_col_email.set(suggest(["email", "e-mail", "mail"]))
        limpeza_col_cnpj.set(suggest(["cnpj"]))

        log_limpeza(f"✅ Colunas carregadas: {len(cols)}")
        messagebox.showinfo("OK", "Colunas carregadas! Agora selecione Razão/Telefones/E-mail/CNPJ.")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao escanear colunas.\n\n{e}")

# ---------------- (NOVO) gráfico embutido na aba limpeza ----------------

graf_canvas = None
graf_fig = None

def _clear_graph():
    global graf_canvas, graf_fig
    if graf_canvas is not None:
        try:
            graf_canvas.get_tk_widget().destroy()
        except:
            pass
    graf_canvas = None
    graf_fig = None

def render_graphs_in_limpeza(counts_reasons: Dict[str, int], counts_uf: Dict[str, int]):
    """
    Desenha 2 gráficos na área direita da aba:
      - Exclusões por motivo
      - Distribuição por UF (derivada do DDD)
    """
    global graf_canvas, graf_fig
    _clear_graph()

    graf_fig = Figure(figsize=(7.5, 5.2), dpi=100)
    ax1 = graf_fig.add_subplot(211)
    ax2 = graf_fig.add_subplot(212)

    # Exclusões por motivo
    if counts_reasons:
        labels = list(counts_reasons.keys())
        values = [counts_reasons[k] for k in labels]
        ax1.bar(labels, values)
        ax1.set_title("Excluídos por motivo")
        ax1.set_ylabel("Qtd")
        ax1.tick_params(axis="x", rotation=25)
    else:
        ax1.text(0.5, 0.5, "Sem exclusões (motivos)", ha="center", va="center")
        ax1.set_axis_off()

    # Por UF (a partir do DDD)
    if counts_uf:
        labels2 = list(counts_uf.keys())
        values2 = [counts_uf[k] for k in labels2]
        ax2.bar(labels2, values2)
        ax2.set_title("Distribuição por UF (via DDD)")
        ax2.set_ylabel("Qtd")
        ax2.tick_params(axis="x", rotation=0)
    else:
        ax2.text(0.5, 0.5, "Sem dados de UF", ha="center", va="center")
        ax2.set_axis_off()

    graf_fig.tight_layout()

    graf_canvas = FigureCanvasTkAgg(graf_fig, master=frame_graficos_limpeza)
    graf_canvas.draw()
    graf_canvas.get_tk_widget().pack(fill="both", expand=True)

# ---------------- (NOVO) deduplicação robusta ----------------

def mark_and_exclude_duplicate_phones(df_base: pd.DataFrame, *, strip55: bool = False, add9: bool = False, add55: bool = False, reason_col: str = "Motivo Exclusao") -> pd.Series:
    """
    Regra (conforme pedido): se um número aparecer mais de uma vez na planilha,
    excluir TODAS as ocorrências, mantendo apenas 1 (a primeira).
    Considera Telefone1 e Telefone2 juntos (como um universo único).
    Retorna máscara booleana de exclusão por duplicidade e preenche o reason_col.
    """
    df = df_base

    # lista "long" com (row_index, phone)
    rows = []
    for idx, (t1, t2) in enumerate(zip(df["Telefone1"].astype(str), df["Telefone2"].astype(str))):
        i = df.index[idx]
        p1 = normalize_phone(t1, strip55=strip55, add9=add9, add55=add55)
        p2 = normalize_phone(t2, strip55=strip55, add9=add9, add55=add55)
        if p1 and not is_invalid_phone(p1):
            rows.append((i, p1))
        if p2 and not is_invalid_phone(p2):
            rows.append((i, p2))

    if not rows:
        return pd.Series(False, index=df.index)

    long = pd.DataFrame(rows, columns=["row", "phone"])
    counts = long["phone"].value_counts()

    # phones duplicados: count > 1
    dup_phones = set(counts[counts > 1].index.tolist())
    if not dup_phones:
        return pd.Series(False, index=df.index)

    # manter a primeira ocorrência global por phone (na ordem em que aparece)
    long["is_dup_phone"] = long["phone"].isin(dup_phones)
    dup_long = long[long["is_dup_phone"]].copy()

    # first occurrence row per phone:
    first_rows = dup_long.drop_duplicates(subset=["phone"], keep="first")[["phone", "row"]]
    keep_pairs = set(zip(first_rows["row"], first_rows["phone"]))

    # tudo que é duplicado e NÃO é a primeira ocorrência -> excluir
    def should_exclude(row, phone):
        return (row, phone) not in keep_pairs

    dup_long["exclude"] = dup_long.apply(lambda r: should_exclude(r["row"], r["phone"]), axis=1)
    rows_to_exclude = set(dup_long.loc[dup_long["exclude"], "row"].tolist())

    mask_dup = df.index.isin(rows_to_exclude)
    if reason_col in df.columns:
        for ridx in df.index[mask_dup]:
            df.at[ridx, reason_col] = _append_reason(df.at[ridx, reason_col], "Telefone duplicado")
    return pd.Series(mask_dup, index=df.index)

def executar_limpeza_dados():
    """
    Nova lógica:
    - Usuário escolhe colunas (Razão, Telefones, Email, CNPJ)
    - Aplica modo de limpeza em Razão Social
    - Divide telefones em Telefone1/Telefone2
    - Remove inválidos (mantém a regra de qualidade)
    - Remove telefones duplicados (mantém apenas 1 ocorrência)
    - Aplica filtros (blocklist c6 + nao perturbe 1..4)
    - Motivo de exclusão por linha (arquivo excluídas)
    - Gráficos embutidos na aba
    - Gera 2 arquivos: filtradas + excluídas
    """
    try:
        txt_log_limpeza.delete("1.0", tk.END)
        progress_limpeza["value"] = 0
        janela.update_idletasks()

        in_path = base_empresas_path.get().strip()
        if not in_path:
            messagebox.showwarning("Aviso", 'Selecione o arquivo "empresas bruto".')
            return

        out_dir = out_dir_limpeza.get().strip() or os.path.dirname(in_path)
        os.makedirs(out_dir, exist_ok=True)

        col_razao = limpeza_col_razao.get().strip()
        col_tel = limpeza_col_tel.get().strip()
        col_email = limpeza_col_email.get().strip()
        col_cnpj = limpeza_col_cnpj.get().strip()

        # Opções de tratamento de telefone (55 / dígito 9 / prefixo 55)
        telefones_tem_55 = (tel_has55_var.get().strip().lower() == "sim")
        opt_strip55 = bool(telefones_tem_55)
        opt_add9 = bool(add9_var.get())
        opt_add55 = bool(add55_var.get())

        if not (col_razao and col_tel and col_email and col_cnpj):
            messagebox.showwarning("Aviso", "Selecione as colunas: Razão Social, Telefones, E-mail e CNPJ.\n\nUse 'Escanear colunas' primeiro.")
            return

        set_status("Executando limpeza de dados...")
        log_limpeza("=== Automação: Limpeza de dados ===")
        log_limpeza(f"📄 Arquivo: {in_path}")
        log_limpeza(f"📁 Saída: {out_dir}")
        log_limpeza(f"🧩 Colunas: Razão='{col_razao}' | Telefones='{col_tel}' | Email='{col_email}' | CNPJ='{col_cnpj}'\n")

        progress_limpeza["value"] = 5
        janela.update_idletasks()

        log_limpeza("1) Lendo arquivo base...")
        df_raw = read_table(in_path)
        log_limpeza(f"✅ Lido: {len(df_raw)} linhas / {len(df_raw.columns)} colunas.")
        progress_limpeza["value"] = 15
        janela.update_idletasks()

        # Monta base só com as colunas escolhidas
        log_limpeza("2) Montando base com as colunas selecionadas...")
        df_base = pd.DataFrame(index=df_raw.index)
        df_base["Razao Social"] = df_raw[col_razao].astype(str)
        df_base["Telefones"] = df_raw[col_tel].astype(str)
        df_base["E-mail"] = df_raw[col_email].astype(str)
        df_base["Cnpj"] = df_raw[col_cnpj].apply(normalize_cnpj)

        # Motivo de exclusão por linha
        df_base["Motivo Exclusao"] = ""

        # Modo de limpeza
        modo = clean_mode_var.get()
        log_limpeza(f"3) Aplicando modo de limpeza na Razão Social: {modo}...")
        if modo == "Lemit":
            df_base["Razao Social"] = df_base["Razao Social"].apply(clean_razao_social)
        else:
            df_base["Razao Social"] = df_base["Razao Social"].astype(str)

        progress_limpeza["value"] = 30
        janela.update_idletasks()

        # Telefones
        log_limpeza("4) Separando e normalizando telefones...")
        t1, t2 = zip(*df_base["Telefones"].map(split_telefones_field))
        df_base["Telefone1"] = pd.Series(t1, index=df_base.index).apply(lambda x: normalize_phone(x, strip55=opt_strip55, add9=opt_add9, add55=opt_add55))
        df_base["Telefone2"] = pd.Series(t2, index=df_base.index).apply(lambda x: normalize_phone(x, strip55=opt_strip55, add9=opt_add9, add55=opt_add55))

        progress_limpeza["value"] = 45
        janela.update_idletasks()

        # Máscara de exclusão
        mask_excluir = pd.Series(False, index=df_base.index)

        # 5) Telefones inválidos
        log_limpeza("5) Removendo linhas sem nenhum telefone válido...")
        invalid_both = df_base["Telefone1"].apply(is_invalid_phone) & df_base["Telefone2"].apply(is_invalid_phone)
        removidas_invalid = int(invalid_both.sum())
        mask_excluir = mask_excluir | invalid_both
        if removidas_invalid:
            df_base.loc[invalid_both, "Motivo Exclusao"] = df_base.loc[invalid_both, "Motivo Exclusao"].apply(lambda s: _append_reason(s, "Telefone inválido"))
        log_limpeza(f"⚠️ Removidas por telefone inválido: {removidas_invalid}")

        progress_limpeza["value"] = 55
        janela.update_idletasks()

        # 6) Duplicados (após inválidos, para não “poluir” contagem)
        log_limpeza("6) Removendo telefones duplicados (mantém apenas 1 ocorrência)...")
        # Só marca duplicados nas linhas ainda não excluídas por inválido
        df_work = df_base[~mask_excluir].copy()
        dup_mask_work = mark_and_exclude_duplicate_phones(df_work, strip55=opt_strip55, add9=opt_add9, add55=opt_add55, reason_col="Motivo Exclusao")
        # mapear de volta pro df_base
        rows_dup = df_work.index[dup_mask_work].tolist()
        dup_mask = df_base.index.isin(rows_dup)
        removidas_dup = int(dup_mask.sum())
        mask_excluir = mask_excluir | dup_mask
        if removidas_dup:
            # df_work já escreveu Motivo Exclusao; precisamos sincronizar
            df_base.loc[rows_dup, "Motivo Exclusao"] = df_work.loc[rows_dup, "Motivo Exclusao"].values
        log_limpeza(f"⚠️ Removidas por duplicidade: {removidas_dup}")

        progress_limpeza["value"] = 60
        janela.update_idletasks()

        # 7) Carrega filtros (Blocklist + Não Perturbe)
        log_limpeza("\n7) Carregando filtros (Blocklist + Não Perturbe 1-4)...")
        filtro_set = set()
        filtro_set |= _load_phone_set(blocklist_c6_path.get(), "Blocklist C6", strip55=opt_strip55)
        filtro_set |= _load_phone_set(nao_perturbe_1_path.get(), "Não Perturbe 1", strip55=opt_strip55)
        filtro_set |= _load_phone_set(nao_perturbe_2_path.get(), "Não Perturbe 2", strip55=opt_strip55)
        filtro_set |= _load_phone_set(nao_perturbe_3_path.get(), "Não Perturbe 3", strip55=opt_strip55)
        filtro_set |= _load_phone_set(nao_perturbe_4_path.get(), "Não Perturbe 4", strip55=opt_strip55)

        progress_limpeza["value"] = 70
        janela.update_idletasks()

        # 8) Aplica filtros por telefone (somente em linhas ainda válidas)
        log_limpeza("8) Aplicando filtros por telefone...")
        if filtro_set:
            candidate = df_base[~mask_excluir]
            in_filters = candidate["Telefone1"].isin(list(filtro_set)) | candidate["Telefone2"].isin(list(filtro_set))
            rows_filter = candidate.index[in_filters].tolist()
            removidas_filtros = len(rows_filter)
            mask_excluir = mask_excluir | df_base.index.isin(rows_filter)
            if removidas_filtros:
                df_base.loc[rows_filter, "Motivo Exclusao"] = df_base.loc[rows_filter, "Motivo Exclusao"].apply(lambda s: _append_reason(s, "Blocklist/Não Perturbe"))
            log_limpeza(f"⚠️ Removidas por Blocklist/Não Perturbe: {removidas_filtros}")
        else:
            log_limpeza("ℹ️ Nenhuma lista selecionada. (Nenhum filtro aplicado)")

        progress_limpeza["value"] = 82
        janela.update_idletasks()

        # Separa finais
        df_excluidas = df_base[mask_excluir].copy()
        df_ficaram = df_base[~mask_excluir].copy()

        # (NOVO) adiciona UF para análise/gráfico (não precisa sair nos arquivos se você não quiser)
        df_ficaram["UF (DDD)"] = df_ficaram.apply(lambda r: uf_from_phone(r.get("Telefone1", ""), r.get("Telefone2", "")), axis=1)
        df_excluidas["UF (DDD)"] = df_excluidas.apply(lambda r: uf_from_phone(r.get("Telefone1", ""), r.get("Telefone2", "")), axis=1)

        log_limpeza("\n9) Preparando arquivos finais (2 resultados)...")
        log_limpeza(f"✅ Ficaram: {len(df_ficaram)}")
        log_limpeza(f"✅ Excluídas: {len(df_excluidas)}")

        # Colunas de saída
        cols_out_filtradas = ["Razao Social", "Telefone1", "Telefone2", "Cnpj", "E-mail"]
        cols_out_excluidas = ["Razao Social", "Telefone1", "Telefone2", "Cnpj", "E-mail", "Motivo Exclusao"]

        for c in cols_out_filtradas:
            if c not in df_ficaram.columns:
                df_ficaram[c] = ""
        for c in cols_out_excluidas:
            if c not in df_excluidas.columns:
                df_excluidas[c] = ""

        df_ficaram_out = df_ficaram[cols_out_filtradas].copy()
        df_excluidas_out = df_excluidas[cols_out_excluidas].copy()

        out_filtradas = os.path.join(out_dir, "empresas_filtradas.xlsx")
        out_excluidas = os.path.join(out_dir, "empresas_excluidas.xlsx")

        safe_remove_file(out_filtradas)
        safe_remove_file(out_excluidas)

        save_to_excel(df_ficaram_out, out_filtradas)
        save_to_excel(df_excluidas_out, out_excluidas)

        progress_limpeza["value"] = 100
        janela.update_idletasks()

        # ----------------- (NOVO) preparar dados dos gráficos -----------------
        # Excluídos por motivo (explode no separador " | ")
        reasons = df_excluidas["Motivo Exclusao"].fillna("").astype(str)
        reason_counts: Dict[str, int] = {}
        for s in reasons.tolist():
            parts = [p.strip() for p in s.split(" | ") if p.strip()]
            if not parts:
                parts = ["(sem motivo)"]
            for p in parts:
                reason_counts[p] = reason_counts.get(p, 0) + 1

        # Distribuição por UF usando DDD (de preferência da base FILTRADA)
        uf_counts_series = df_ficaram["UF (DDD)"].value_counts()
        uf_counts: Dict[str, int] = uf_counts_series.to_dict()

        # Renderiza os gráficos na aba
        render_graphs_in_limpeza(reason_counts, uf_counts)

        log_limpeza("\n🎉 Processo concluído!")
        log_limpeza(f"📄 Gerado: {out_filtradas}")
        log_limpeza(f"📄 Gerado: {out_excluidas}")

        messagebox.showinfo(
            "Concluído",
            "Limpeza finalizada!\n\n"
            f"Filtradas: {os.path.basename(out_filtradas)}\n"
            f"Excluídas: {os.path.basename(out_excluidas)}\n\n"
            "Obs: O arquivo de excluídas contém a coluna 'Motivo Exclusao'."
        )

    except Exception as e:
        log_limpeza(f"\n❌ Erro fatal: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro durante a execução.\n\n{e}")


# =======================================================================
#           FUNÇÕES ROBÔ C6
# =======================================================================

def log_robo(msg: str):
    txt_log_robo.insert(tk.END, msg + "\n")
    txt_log_robo.see(tk.END)
    janela.update_idletasks()

def selecionar_arquivos_robo():
    paths = filedialog.askopenfilenames(
        title="Selecione as planilhas (até 20 mil linhas cada)",
        filetypes=[("Planilhas", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if paths:
        robo_arquivos.clear()
        robo_arquivos.extend(paths)
        lbl_arquivos_robo.config(text=f"{len(robo_arquivos)} arquivo(s) selecionado(s).")

def selecionar_bat_robo():
    path = filedialog.askopenfilename(
        title="Selecione o arquivo .BAT",
        filetypes=[("Arquivos BAT", "*.bat"), ("Todos os arquivos", "*.*")]
    )
    if path:
        robo_bat_path.set(path)
        lbl_bat_robo.config(text=f".BAT selecionado: {path}")
        bat_dir = os.path.dirname(path)
        resultado_sugerido = os.path.join(bat_dir, "resultado")
        if os.path.isdir(resultado_sugerido) and not robo_resultado_dir.get():
            robo_resultado_dir.set(resultado_sugerido)
            lbl_pasta_resultado.config(text=f"Pasta de resultados: {resultado_sugerido}")

def selecionar_resultado_robo():
    path = filedialog.askdirectory(title="Selecione a pasta de resultados do .BAT")
    if path:
        robo_resultado_dir.set(path)
        lbl_pasta_resultado.config(text=f"Pasta de resultados: {path}")

def executar_robo_c6():
    try:
        txt_log_robo.delete("1.0", tk.END)
        progress_robo["value"] = 0
        janela.update_idletasks()

        if not robo_arquivos:
            messagebox.showwarning("Aviso", "Selecione as planilhas de entrada primeiro.")
            return

        bat_path = robo_bat_path.get().strip()
        if not bat_path or not os.path.isfile(bat_path):
            messagebox.showwarning("Aviso", "Selecione um arquivo .BAT válido.")
            return

        resultado_dir = robo_resultado_dir.get().strip()
        if not resultado_dir or not os.path.isdir(resultado_dir):
            messagebox.showwarning("Aviso", "Selecione uma pasta de resultados válida do .BAT.")
            return

        modo = robo_modo_var.get()
        if modo not in ["Lemit", "Simples"]:
            messagebox.showwarning("Aviso", "Selecione se o arquivo é para Lemit ou Simples.")
            return

        bat_dir = os.path.dirname(bat_path)
        log_robo("=== Robô C6 iniciado ===")
        log_robo(f"Arquivos selecionados: {len(robo_arquivos)}")
        log_robo(f"Caminho do .BAT: {bat_path}")
        log_robo(f"Pasta de resultados do .BAT: {resultado_dir}")
        log_robo(f"Modo de tratamento final: {modo}\n")

        total_arquivos = len(robo_arquivos)

        log_robo("Limpando pasta de resultados antes de iniciar...")
        for f in os.listdir(resultado_dir):
            full = os.path.join(resultado_dir, f)
            if os.path.isfile(full) and any(full.lower().endswith(ext) for ext in [".xlsx", ".xls", ".csv", ".txt"]):
                safe_remove_file(full)
        log_robo("Pasta de resultados limpa.\n")

        for idx, arquivo in enumerate(robo_arquivos, start=1):
            progress_robo["value"] = (idx - 1) / total_arquivos * 40
            janela.update_idletasks()

            log_robo(f"[{idx}/{total_arquivos}] Preparando arquivo: {arquivo}")
            try:
                base_name = os.path.basename(arquivo)
                dest_path = os.path.join(bat_dir, base_name)
                shutil.copy2(arquivo, dest_path)
                log_robo(f"→ Copiado para pasta do .BAT: {dest_path}")
            except Exception as e:
                log_robo(f"❌ Erro ao copiar arquivo para pasta do .BAT: {e}")
                continue

            try:
                log_robo("→ Executando .BAT...")
                proc = subprocess.Popen(
                    bat_path,
                    cwd=bat_dir,
                    stdin=subprocess.PIPE,
                    shell=True
                )
                proc.communicate(input=b"\n")
                log_robo("→ Execução do .BAT concluída.")
            except Exception as e:
                log_robo(f"❌ Erro ao executar .BAT: {e}")
                continue

            if idx < total_arquivos:
                log_robo("⏱ Aguardando 6 minutos antes do próximo arquivo...")
                janela.update_idletasks()
                time.sleep(6 * 60)
                log_robo("✔ Intervalo concluído.\n")
            else:
                log_robo("Último arquivo processado.\n")

            progress_robo["value"] = idx / total_arquivos * 60
            janela.update_idletasks()

        log_robo("Lendo arquivos de resultados gerados pelo .BAT...")
        result_files = []
        for f in os.listdir(resultado_dir):
            full = os.path.join(resultado_dir, f)
            if os.path.isfile(full) and full.lower().endswith((".xlsx", ".xls", ".csv", ".txt")):
                result_files.append(full)

        if not result_files:
            log_robo("⚠️ Nenhum arquivo de resultado encontrado na pasta informada.")
            messagebox.showwarning("Aviso", "Nenhum arquivo de resultado foi encontrado na pasta de resultados.")
            progress_robo["value"] = 100
            return

        log_robo(f"Encontrados {len(result_files)} arquivo(s) de resultado.")
        dfs = []
        for fpath in result_files:
            try:
                log_robo(f"Lendo resultado: {fpath}")
                dfs.append(read_table(fpath))
            except Exception as e:
                log_robo(f"❌ Erro ao ler resultado {fpath}: {e}")

        if not dfs:
            log_robo("⚠️ Não foi possível ler nenhum arquivo de resultado.")
            progress_robo["value"] = 100
            return

        df_total = pd.concat(dfs, ignore_index=True)
        log_robo(f"Total de linhas combinadas (antes da filtragem): {len(df_total)}")

        log_robo("Aplicando filtro: remover linhas com 'Nao disponivel' e manter apenas 'Novo cliente'...")
        df_str = df_total.astype(str)

        mask_nao_disponivel = df_str.apply(
            lambda col: col.str.contains("Nao disponivel", case=False, na=False)
        ).any(axis=1)

        mask_novo_cliente = df_str.apply(
            lambda col: col.str.contains("Novo cliente", case=False, na=False)
        ).any(axis=1)

        antes = len(df_total)
        df_filtrado = df_total[~mask_nao_disponivel & mask_novo_cliente].copy()
        removidas = antes - len(df_filtrado)
        log_robo(f"Linhas removidas pelo filtro: {removidas}")
        log_robo(f"Linhas finais após filtro: {len(df_filtrado)}")

        if df_filtrado.empty:
            log_robo("⚠️ Nenhuma linha restante após aplicar o filtro.")
            messagebox.showinfo("Concluído", "Robô C6 finalizado, mas nenhuma linha restou após o filtro.")
            progress_robo["value"] = 100
            return

        progress_robo["value"] = 80
        janela.update_idletasks()

        if modo == "Lemit":
            log_robo("Modo Lemit: gerando 1 planilha única com o resultado final...")
            out_path = os.path.join(resultado_dir, "robo_c6_final_LEMIT.xlsx")
            save_to_excel(df_filtrado, out_path)
            log_robo(f"✅ Arquivo final gerado: {out_path}")
        else:
            log_robo("Modo Simples: separando resultado em planilhas de 5.000 linhas...")
            total = len(df_filtrado)
            chunk_size = 5000
            parts = math.ceil(total / chunk_size)
            log_robo(f"Total de linhas: {total} → {parts} arquivo(s) de até {chunk_size} linhas.")

            for i in range(parts):
                start = i * chunk_size
                end = min(start + chunk_size, total)
                part = df_filtrado.iloc[start:end]
                out_path = os.path.join(resultado_dir, f"robo_c6_SIMPLES_part{i+1}.xlsx")
                save_to_excel(part, out_path)
                log_robo(f"✅ Parte {i+1} salva: {out_path} ({len(part)} linhas)")

        progress_robo["value"] = 100
        janela.update_idletasks()
        log_robo("\n🎉 Robô C6 concluído com sucesso!")
        messagebox.showinfo("Concluído", "Robô C6 finalizado com sucesso!")

    except Exception as e:
        log_robo(f"\n❌ Erro fatal no Robô C6: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro durante o Robô C6.\n\n{e}")


# =======================================================================
#           ABA MANIPULAÇÃO (NOVA)
# =======================================================================

def log_manip(msg: str):
    txt_log_manip.insert(tk.END, msg + "\n")
    txt_log_manip.see(tk.END)
    janela.update_idletasks()

def selecionar_arquivos_manip():
    paths = filedialog.askopenfilenames(
        title="Selecione as planilhas",
        filetypes=[("Planilhas", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if paths:
        manip_arquivos.clear()
        manip_arquivos.extend(paths)
        lbl_manip_arquivos.config(text=f"{len(manip_arquivos)} arquivo(s) selecionado(s).")

def selecionar_out_dir_manip():
    path = filedialog.askdirectory(title="Selecione a pasta de saída")
    if path:
        manip_out_dir.set(path)
        lbl_manip_saida.config(text=f"Saída: {path}")

def executar_manipulacao():
    try:
        txt_log_manip.delete("1.0", tk.END)

        if not manip_arquivos:
            messagebox.showwarning("Aviso", "Selecione as planilhas primeiro.")
            return

        out_dir = manip_out_dir.get().strip()
        if not out_dir:
            messagebox.showwarning("Aviso", "Selecione a pasta de saída.")
            return

        os.makedirs(out_dir, exist_ok=True)

        modo = manip_modo_var.get()
        log_manip("=== Manipulação de planilhas ===")
        log_manip(f"Modo: {modo}")
        log_manip(f"Arquivos: {len(manip_arquivos)}")
        log_manip(f"Saída: {out_dir}\n")

        log_manip("1) Lendo planilhas...")
        dfs = []
        for f in manip_arquivos:
            log_manip(f"→ Lendo: {f}")
            dfs.append(read_table(f))

        df_all = pd.concat(dfs, ignore_index=True)
        log_manip(f"✅ Total combinado: {len(df_all)} linhas.\n")

        if modo == "juntar":
            out_path = os.path.join(out_dir, "planilhas_juntas.xlsx")
            safe_remove_file(out_path)
            log_manip("2) Salvando planilha única...")
            save_to_excel(df_all, out_path)
            log_manip(f"✅ Gerado: {out_path}")
            messagebox.showinfo("Concluído", f"Planilhas juntadas com sucesso!\n\n{out_path}")
            return

        # separar
        try:
            chunk = int(manip_linhas_por_planilha.get().strip())
            if chunk <= 0:
                raise ValueError()
        except:
            messagebox.showerror("Erro", "Quantidade de linhas por planilha inválida.")
            return

        total = len(df_all)
        parts = math.ceil(total / chunk)
        log_manip(f"2) Separando em partes de {chunk} linhas → {parts} arquivos...")

        for i in range(parts):
            start = i * chunk
            end = min(start + chunk, total)
            part = df_all.iloc[start:end]
            out_path = os.path.join(out_dir, f"separado_part{i+1}.xlsx")
            safe_remove_file(out_path)
            save_to_excel(part, out_path)
            log_manip(f"✅ Parte {i+1}: {out_path} ({len(part)} linhas)")

        messagebox.showinfo("Concluído", f"Separação concluída! Gerados {parts} arquivo(s) em:\n\n{out_dir}")

    except Exception as e:
        log_manip(f"\n❌ Erro: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro na manipulação.\n\n{e}")


# =======================================================================
#           INTERFACE GRÁFICA (TKINTER)
# =======================================================================

janela = tk.Tk()
janela.title("B2BSAFE")
janela.geometry("1350x780")
janela.configure(bg=BG_PRINCIPAL)


# -------------------- BARRA SUPERIOR + STATUS --------------------
APP_NOME = "B2BSAFE"
APP_VERSAO = "v5.1 (UI Pro)"

def _hex_to_rgb(hexcolor: str):
    hexcolor = hexcolor.lstrip("#")
    return tuple(int(hexcolor[i:i+2], 16) for i in (0, 2, 4))

def _blend(hex_a: str, hex_b: str, t: float):
    a = _hex_to_rgb(hex_a); b = _hex_to_rgb(hex_b)
    c = tuple(int(a[i]*(1-t)+b[i]*t) for i in range(3))
    return "#%02x%02x%02x" % c

status_var = tk.StringVar(value="Pronto.")

# (UI) Cabeçalho removido: o branding fica na sidebar.

style = ttk.Style()
try:
    style.theme_use("clam")
except:
    pass


# ---- Estilo do Notebook (abas) / Treeview (visual mais empresarial) ----
style.configure("TNotebook", background=BG_PRINCIPAL, borderwidth=0)
style.configure("TNotebook.Tab",
                background=_blend(BG_FRAME, BG_PRINCIPAL, 0.25),
                foreground=FG_SECUNDARIO,
                padding=(14, 8),
                font=("Segoe UI", 10, "bold"),
                borderwidth=0)
style.map("TNotebook.Tab",
          background=[("selected", BG_FRAME), ("active", _blend(BG_FRAME, BG_PRINCIPAL, 0.10))],
          foreground=[("selected", FG_TEXTO), ("active", FG_TEXTO)])

style.configure("Treeview",
                background=INPUT_BG,
                fieldbackground=INPUT_BG,
                foreground=INPUT_FG,
                bordercolor=BORDER_COR,
                rowheight=26)
style.configure("Treeview.Heading",
                background=BG_FRAME,
                foreground=FG_TEXTO,
                font=("Segoe UI", 10, "bold"))
style.map("Treeview", background=[("selected", _blend(ACCENT_BLUE, BG_PRINCIPAL, 0.25))])

style.configure("Vertical.TScrollbar",
                troughcolor=BG_FRAME,
                background=_blend(BORDER_COR, BG_FRAME, 0.25),
                bordercolor=BG_FRAME,
                arrowcolor=FG_TEXTO)
# -----------------------------------------------------------------------
style.configure("TLabel", background=BG_PRINCIPAL, foreground=FG_TEXTO, font=fonte_label)

style.configure("Frame.TLabelframe",
                background=BG_FRAME,
                foreground=FG_TEXTO,
                bordercolor=BORDER_COR)
style.configure("Frame.TLabelframe.Label",
                background=BG_FRAME,
                foreground=FG_TEXTO,
                font=("Segoe UI", 11, "bold"))

style.configure(
    "Custom.Horizontal.TProgressbar",
    troughcolor=BG_FRAME,
    background=ACCENT_GREEN,
    bordercolor=BG_FRAME,
    lightcolor=ACCENT_GREEN,
    darkcolor=ACCENT_GREEN
)

style.configure("TCombobox",
                fieldbackground=INPUT_BG,
                background=INPUT_BG,
                foreground=INPUT_FG,
                arrowcolor=INPUT_FG,
                bordercolor=BORDER_COR)
style.map("TCombobox",
          fieldbackground=[("readonly", INPUT_BG)],
          foreground=[("readonly", INPUT_FG)])

style.configure("Accent.TButton", font=("Segoe UI Semibold", 10), padding=10,
    foreground="white",
    background=ACCENT_GREEN,
    borderwidth=0,
    focuscolor=BG_PRINCIPAL
)
style.map("Accent.TButton",
          background=[("active", "#16A34A")],
          foreground=[("active", "white")])

style.configure("Primary.TButton", font=("Segoe UI Semibold", 10), padding=10,
    foreground="white",
    background=ACCENT_BLUE,
    borderwidth=0,
    focuscolor=BG_PRINCIPAL
)
style.map("Primary.TButton",
          background=[("active", "#1D4ED8")],
          foreground=[("active", "white")])

style.configure("Warn.TButton", font=("Segoe UI Semibold", 10), padding=10,
    foreground="white",
    background=ACCENT_ORANGE,
    borderwidth=0,
    focuscolor=BG_PRINCIPAL
)
style.map("Warn.TButton",
          background=[("active", "#EA580C")],
          foreground=[("active", "white")])

# -------------------- UI HELPERS (SaaS feel) --------------------
def _bind_hover(ttk_btn: ttk.Widget):
    try:
        ttk_btn.configure(cursor="hand2")
    except Exception:
        pass

def _apply_hover_style(btn: ttk.Button, normal_style: str, hover_style: str):
    def on_enter(_e):
        try:
            btn.configure(style=hover_style)
        except Exception:
            pass
    def on_leave(_e):
        try:
            btn.configure(style=normal_style)
        except Exception:
            pass
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    _bind_hover(btn)

# Hover styles (same base, apenas muda cor via 'active' map)
style.configure("AccentHover.TButton", **style.configure("Accent.TButton"))
style.map("AccentHover.TButton", background=[("active", "#15803D"), ("!disabled", "#16A34A")])

style.configure("PrimaryHover.TButton", **style.configure("Primary.TButton"))
style.map("PrimaryHover.TButton", background=[("active", "#1E40AF"), ("!disabled", "#2563EB")])

style.configure("WarnHover.TButton", **style.configure("Warn.TButton"))
style.map("WarnHover.TButton", background=[("active", "#C2410C"), ("!disabled", "#F97316")])




# =======================================================================
#           LAYOUT EMPRESARIAL: SIDEBAR + CONTEÚDO (COM SCROLL)
# =======================================================================

class ScrollableFrame(tk.Frame):
    def __init__(self, parent, bg, *args, **kwargs):
        super().__init__(parent, bg=bg, *args, **kwargs)

        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0)
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set)

        self.inner = tk.Frame(self.canvas, bg=bg)
        self.inner_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.pack(side="left", fill="both", expand=True)
        self.v_scroll.pack(side="right", fill="y")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Mousewheel (Windows/Mac/Linux)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)       # Windows / Mac
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)   # Linux up
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)   # Linux down

    def _on_inner_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # manter a largura do inner igual à do canvas
        self.canvas.itemconfigure(self.inner_id, width=event.width)

    def _on_mousewheel(self, event):
        # Evita scroll quando o widget não está visível
        if not self.winfo_ismapped():
            return
        delta = -1 * int(event.delta / 120) if event.delta else 0
        self.canvas.yview_scroll(delta, "units")

    def _on_mousewheel_linux(self, event):
        if not self.winfo_ismapped():
            return
        if event.num == 4:
            self.canvas.yview_scroll(-3, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(3, "units")


# -------------------- Temas (Claro/Escuro) --------------------
THEMES = {
    # Inspirado em SaaS modernos (Linear/Notion/Stripe): alto contraste, superfícies bem definidas e acento consistente.
    "Escuro": {
        "BG_PRINCIPAL": "#0B0F19",   # fundo
        "BG_FRAME": "#111827",       # cards/superfícies
        "FG_TEXTO": "#E5E7EB",
        "FG_SECUNDARIO": "#9CA3AF",
        "ACCENT_GREEN": "#22C55E",
        "ACCENT_BLUE": "#4F46E5",    # indigo (mais “enterprise”)
        "ACCENT_ORANGE": "#F97316",
        "BORDER_COR": "#1F2937",
        "INPUT_BG": "#0F172A",
        "INPUT_FG": "#F9FAFB",
    },
    "Claro": {
        "BG_PRINCIPAL": "#F8FAFC",
        "BG_FRAME": "#FFFFFF",
        "FG_TEXTO": "#0F172A",
        "FG_SECUNDARIO": "#475569",
        "ACCENT_GREEN": "#16A34A",
        "ACCENT_BLUE": "#4F46E5",
        "ACCENT_ORANGE": "#EA580C",
        "BORDER_COR": "#E2E8F0",
        "INPUT_BG": "#FFFFFF",
        "INPUT_FG": "#0F172A",
    }
}

current_theme_name = tk.StringVar(value="Escuro")

def _apply_theme_constants(theme_name: str):
    # Atualiza as constantes globais usadas no resto do app
    global BG_PRINCIPAL, BG_FRAME, FG_TEXTO, FG_SECUNDARIO, ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE, BORDER_COR, INPUT_BG, INPUT_FG
    t = THEMES[theme_name]
    BG_PRINCIPAL = t["BG_PRINCIPAL"]
    BG_FRAME = t["BG_FRAME"]
    FG_TEXTO = t["FG_TEXTO"]
    FG_SECUNDARIO = t["FG_SECUNDARIO"]
    ACCENT_GREEN = t["ACCENT_GREEN"]
    ACCENT_BLUE = t["ACCENT_BLUE"]
    ACCENT_ORANGE = t["ACCENT_ORANGE"]
    BORDER_COR = t["BORDER_COR"]
    INPUT_BG = t["INPUT_BG"]
    INPUT_FG = t["INPUT_FG"]

def _restyle_ttk():
    # Reaplica os estilos ttk para o tema atual
    style.configure("TLabel", background=BG_PRINCIPAL, foreground=FG_TEXTO, font=fonte_label)

    style.configure("Frame.TLabelframe", background=BG_FRAME, foreground=FG_TEXTO, bordercolor=BORDER_COR)
    style.configure("Frame.TLabelframe.Label", background=BG_FRAME, foreground=FG_TEXTO, font=("Segoe UI", 11, "bold"))

    style.configure("Custom.Horizontal.TProgressbar",
                    troughcolor=BG_FRAME,
                    background=ACCENT_GREEN,
                    bordercolor=BG_FRAME,
                    lightcolor=ACCENT_GREEN,
                    darkcolor=ACCENT_GREEN)

    style.configure("TCombobox",
                    fieldbackground=INPUT_BG,
                    background=INPUT_BG,
                    foreground=INPUT_FG,
                    arrowcolor=INPUT_FG,
                    bordercolor=BORDER_COR)
    style.map("TCombobox",
              fieldbackground=[("readonly", INPUT_BG)],
              foreground=[("readonly", INPUT_FG)])

    style.configure("Accent.TButton", font=("Segoe UI Semibold", 10), foreground="white", background=ACCENT_GREEN, borderwidth=0, focuscolor=BG_PRINCIPAL)
    style.map("Accent.TButton", background=[("active", "#16A34A") if current_theme_name.get()=="Escuro" else ("active", "#15803D")], foreground=[("active", "white")])

    style.configure("Primary.TButton", font=("Segoe UI Semibold", 10), foreground="white", background=ACCENT_BLUE, borderwidth=0, focuscolor=BG_PRINCIPAL)
    style.map("Primary.TButton", background=[("active", "#1D4ED8")], foreground=[("active", "white")])

    style.configure("Warn.TButton", font=("Segoe UI Semibold", 10), foreground="white", background=ACCENT_ORANGE, borderwidth=0, focuscolor=BG_PRINCIPAL)
    style.map("Warn.TButton", background=[("active", "#EA580C")], foreground=[("active", "white")])

def _restyle_tk_widgets(widget):
    # Atualiza bg/fg de widgets tk.* (ttk usa style)
    try:
        cls = widget.__class__.__name__
        if isinstance(widget, (tk.Frame, tk.Label, tk.Button)):
            if "bg" in widget.configure():
                widget.configure(bg=BG_PRINCIPAL if isinstance(widget, tk.Frame) else widget.cget("bg"))
        if isinstance(widget, tk.Label):
            if "bg" in widget.configure():
                # Mantém labels em BG_PRINCIPAL ou BG_FRAME dependendo do pai
                parent_bg = widget.master.cget("bg") if hasattr(widget.master, "cget") else BG_PRINCIPAL
                widget.configure(bg=parent_bg, fg=FG_TEXTO)
        if isinstance(widget, tk.Text):
            widget.configure(bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
        if isinstance(widget, tk.Entry):
            widget.configure(bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG)
    except Exception:
        pass

    for child in widget.winfo_children():
        _restyle_tk_widgets(child)

def set_theme(theme_name: str):
    """Aplica o tema (Claro/Escuro) e reestiliza widgets ttk/tk.

    Observação: alguns elementos (sidebar/content) só existem depois de montar a UI,
    então aqui usamos try/except para não quebrar a inicialização.
    """
    _apply_theme_constants(theme_name)
    janela.configure(bg=BG_PRINCIPAL)
    _restyle_ttk()
    _restyle_tk_widgets(janela)

    # Atualiza containers que dependem do tema (se já existirem)
    try:
        main.configure(bg=BG_PRINCIPAL)
        body.configure(bg=BG_PRINCIPAL)
        sidebar.configure(bg=BG_FRAME)
        content.configure(bg=BG_PRINCIPAL)

        # Branding/sidebar
        brand_box.configure(bg=BG_FRAME)
        brand_label.configure(bg=BG_FRAME, fg=FG_TEXTO)
        brand_sub.configure(bg=BG_FRAME, fg=FG_SECUNDARIO)
        title_nav.configure(bg=BG_FRAME, fg=FG_SECUNDARIO)

        theme_box.configure(bg=BG_FRAME)
        theme_switch.configure(bg=BG_FRAME)
        _draw_theme_switch()

        sidebar_status_label.configure(bg=BG_FRAME, fg=FG_SECUNDARIO)
    except Exception:
        pass

# aplica tema inicial
# (mantém compatibilidade com o resto do arquivo)
_apply_theme_constants(current_theme_name.get())
_restyle_ttk()


# -------------------- Estrutura principal (Sidebar + Conteúdo) --------------------
main = tk.Frame(janela, bg=BG_PRINCIPAL)
main.pack(fill="both", expand=True)

# body
body = tk.Frame(main, bg=BG_PRINCIPAL)
body.pack(side="top", fill="both", expand=True, padx=10, pady=(0, 8))

# Sidebar retrátil
sidebar_expanded = tk.BooleanVar(value=True)
sidebar = tk.Frame(body, bg=BG_FRAME, width=260)
sidebar.pack(side="left", fill="y")
sidebar.pack_propagate(False)

# -------------------- Branding (vai recolher junto com a sidebar) --------------------
brand_box = tk.Frame(sidebar, bg=BG_FRAME)
brand_box.pack(fill="x", padx=12, pady=(12, 8))

brand_label = tk.Label(brand_box, text="B2BSAFE", bg=BG_FRAME, fg=FG_TEXTO, font=("Segoe UI", 14, "bold"))
brand_label.pack(anchor="w")

brand_sub = tk.Label(brand_box, text="Suite • Automação", bg=BG_FRAME, fg=FG_SECUNDARIO, font=("Segoe UI", 9))
brand_sub.pack(anchor="w")

# -------------------- Toggle Tema (switch on/off) --------------------
theme_box = tk.Frame(sidebar, bg=BG_FRAME)
theme_box.pack(fill="x", padx=12, pady=(0, 8))

tk.Label(theme_box, text="Tema", bg=BG_FRAME, fg=FG_SECUNDARIO, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 4))

theme_is_dark = tk.BooleanVar(value=(current_theme_name.get() == "Escuro"))
theme_switch = tk.Canvas(theme_box, width=52, height=26, bg=BG_FRAME, highlightthickness=0)
theme_switch.pack(anchor="w")

def _draw_theme_switch():
    # desenha um switch simples (SaaS-like)
    theme_switch.delete("all")
    on = theme_is_dark.get()
    track = _blend(ACCENT_BLUE, BG_FRAME, 0.15) if on else _blend(FG_SECUNDARIO, BG_FRAME, 0.65)
    knob = "#FFFFFF" if current_theme_name.get() == "Escuro" else "#0F172A"
    theme_switch.create_rounded_rect = getattr(theme_switch, "create_rounded_rect", None)

    # rounded rect manual
    x0, y0, x1, y1, r = 2, 4, 50, 22, 10
    theme_switch.create_rectangle(x0+r, y0, x1-r, y1, fill=track, outline=track)
    theme_switch.create_rectangle(x0, y0+r, x1, y1-r, fill=track, outline=track)
    theme_switch.create_oval(x0, y0, x0+2*r, y0+2*r, fill=track, outline=track)
    theme_switch.create_oval(x1-2*r, y0, x1, y0+2*r, fill=track, outline=track)

    # knob
    cx = 36 if on else 16
    theme_switch.create_oval(cx-9, 13-9, cx+9, 13+9, fill=knob, outline=knob)

    # ícone pequeno
    icon = "🌙" if on else "☀️"
    theme_switch.create_text(8 if on else 44, 13, text=icon, font=("Segoe UI", 9))

def _toggle_theme_from_switch(event=None):
    theme_is_dark.set(not theme_is_dark.get())
    current_theme_name.set("Escuro" if theme_is_dark.get() else "Claro")
    set_theme(current_theme_name.get())

theme_switch.bind("<Button-1>", _toggle_theme_from_switch)
_draw_theme_switch()

# -------------------- Status (fica no menu lateral) --------------------
theme_box.pack(side="bottom", fill="x", padx=12, pady=(0, 8))

sidebar_status_label = tk.Label(sidebar, text="Pronto.", bg=BG_FRAME, fg=FG_SECUNDARIO, font=("Segoe UI", 9))
sidebar_status_label.pack(side="bottom", fill="x", padx=12, pady=12)


content = tk.Frame(body, bg=BG_PRINCIPAL)
content.pack(side="left", fill="both", expand=True, padx=(10, 0))

def toggle_sidebar():
    # Sidebar retrátil: recolhido mostra só ícones; aberto mostra ícone + texto.
    global sidebar_collapsed
    sidebar_collapsed = not sidebar_collapsed

    if sidebar_collapsed:
        sidebar.configure(width=76)
        # Texto vira ícone
        for b, full_text, short_text in sidebar_buttons:
            b.config(text=short_text, anchor="center")
        # Branding compacto
        brand_label.config(text="B2B")
        brand_sub.pack_forget()
        # reduz paddings visuais
        toggle_btn.config(text="☰")
    else:
        sidebar.configure(width=260)
        for b, full_text, short_text in sidebar_buttons:
            b.config(text=full_text, anchor="w")
        brand_label.config(text="B2BSAFE")
        if not brand_sub.winfo_ismapped():
            brand_sub.pack(anchor="w")
        toggle_btn.config(text="☰")

    sidebar.update_idletasks()


toggle_btn = ttk.Button(sidebar, text="☰", style="Primary.TButton", command=toggle_sidebar)
toggle_btn.pack(fill="x", padx=10, pady=(10, 8))

title_nav = tk.Label(sidebar, text="", bg=BG_FRAME, fg=FG_SECUNDARIO, font=("Segoe UI", 9))
# (removido "Menu" no sidebar)

# Área de módulos (frames com scroll)
sf_home = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_procv = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_limpeza = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_wpp = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_robo = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_manip = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_bd = ScrollableFrame(content, bg=BG_PRINCIPAL)
sf_conexao = ScrollableFrame(content, bg=BG_PRINCIPAL)

# Empilha módulos no mesmo espaço (somente o menu fica fixo).
content.grid_rowconfigure(0, weight=1)
content.grid_columnconfigure(0, weight=1)

for sf in [sf_home, sf_procv, sf_limpeza, sf_wpp, sf_robo, sf_manip, sf_bd, sf_conexao]:
    sf.grid(row=0, column=0, sticky="nsew")

frame_home = sf_home.inner
frame_procv = sf_procv.inner
frame_limpeza = sf_limpeza.inner
frame_wpp = sf_wpp.inner
frame_robo = sf_robo.inner
frame_manip = sf_manip.inner
frame_bd = sf_bd.inner
frame_conexao = sf_conexao.inner

frames_map = {
    "home": sf_home,
    "procv": sf_procv,
    "limpeza": sf_limpeza,
    "wpp": sf_wpp,
    "robo": sf_robo,
    "manip": sf_manip,
    "bd": sf_bd,
    "conexao": sf_conexao,
}

# Status (embutido na sidebar)
# set_status será ligado ao label da sidebar após construir o menu.

def set_status(msg: str):
    try:
        sidebar_status_label.config(text=msg)
        janela.update_idletasks()
    except Exception:
        pass

def show_frame(key: str):
    f = frames_map.get(key)
    if not f:
        return

    # troca de página
    f.tkraise()
    set_status("Pronto.")

    # destaque do menu selecionado
    for b, full_text, short_text in sidebar_buttons:
        # compara pela função associada (key no lambda) via texto completo cadastrado em _add_nav
        if getattr(b, "_nav_key", None) == key:
            b.configure(style="NavActive.TButton")
        else:
            b.configure(style="Nav.TButton")

# Botões do sidebar
sidebar_buttons = []
def _add_nav(text, key, short):
    b = ttk.Button(sidebar, text=text, style="Nav.TButton", command=lambda k=key: show_frame(k))
    b._nav_key = key
    b.pack(fill="x", padx=10, pady=6)
    sidebar_buttons.append((b, text, short))

_add_nav("Início", "home", "🏠")
_add_nav("PROCV B2B", "procv", "🔎")
_add_nav("Limpeza de dados", "limpeza", "🧹")
_add_nav("Limpeza WhatsApp", "wpp", "💬")
_add_nav("Robô C6", "robo", "🤖")
_add_nav("Manipulação", "manip", "🧩")
_add_nav("Banco de Dados", "bd", "🗄️")
_add_nav("Conexão BD", "conexao", "🔌")

# Inicial
show_frame("home")



# =======================================================================
#                           ABA INÍCIO
# =======================================================================

lbl_home_title = tk.Label(
    frame_home,
    text="Bem-vindo à sua suíte de ferramentas ⚙️",
    bg=BG_PRINCIPAL,
    fg=FG_TEXTO,
    font=fonte_titulo
)
lbl_home_title.pack(pady=(40, 10))

lbl_home_sub = tk.Label(
    frame_home,
    text=(
        "Aqui você pode centralizar várias automações e utilitários.\n"
        "Módulos disponíveis:\n"
        "• PROCV B2B (comparação de colunas)\n"
        "• Limpeza de dados (tratamento de base + filtros por telefone)\n"
        "• Robô C6 (automação de .BAT + consolidação de resultados)\n"
        "• Manipulação (juntar/separar planilhas)\n"
        "• Banco de Dados (importar arquivos Excel/CSV para tabelas + visualizar)\n"
        "• Conexão BD (configuração e conexão com banco MySQL/PostgreSQL)"
    ),
    bg=BG_PRINCIPAL,
    fg=FG_SECUNDARIO,
    font=("Segoe UI", 11),
    justify="center"
)
lbl_home_sub.pack(pady=(0, 30))

def ir_para(key):
    show_frame(key)
ttk.Button(frame_home, text="Abrir módulo PROCV B2B", style="Accent.TButton", command=lambda: ir_para("procv")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Limpeza de dados", style="Primary.TButton", command=lambda: ir_para("limpeza")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Limpeza WhatsApp", style="Primary.TButton", command=lambda: ir_para("wpp")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Robô C6", style="Warn.TButton", command=lambda: ir_para("robo")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Manipulação", style="Primary.TButton", command=lambda: ir_para("manip")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Banco de Dados", style="Primary.TButton", command=lambda: ir_para("bd")).pack(pady=8)
ttk.Button(frame_home, text="Abrir módulo Conexão BD", style="Primary.TButton", command=lambda: ir_para("conexao")).pack(pady=8)


# =======================================================================
#                       ABA PROCV B2B
# =======================================================================

caminho_arquivo = tk.StringVar()
caminho_arquivo_saida = tk.StringVar()
pasta_saida = tk.StringVar()

lbl_titulo = tk.Label(frame_procv, text="PROCV B2B - Comparador de Colunas", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_titulo.pack(pady=(10, 2))

lbl_sub = tk.Label(frame_procv, text="Compare colunas de arquivos CSV/XLSX e gere um Excel com os itens exclusivos, já destacados.",
                   bg=BG_PRINCIPAL, fg=FG_SECUNDARIO, font=("Segoe UI", 10))
lbl_sub.pack(pady=(0, 10))

frame_arquivo = ttk.Labelframe(frame_procv, text="Arquivo de entrada", style="Frame.TLabelframe", padding=10)
frame_arquivo.pack(padx=12, pady=6, fill="x")

entry_arquivo = tk.Entry(frame_arquivo, textvariable=caminho_arquivo, font=fonte_entry, width=65,
                         bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, bd=1, relief="solid", highlightthickness=0)
entry_arquivo.pack(side=tk.LEFT, padx=5, pady=3)

def selecionar_arquivo():
    arquivo = filedialog.askopenfilename(
        title="Selecione um arquivo",
        filetypes=[("Excel e CSV", "*.xlsx *.csv"), ("Todos os arquivos", "*.*")]
    )
    if arquivo:
        caminho_arquivo.set(arquivo)

ttk.Button(frame_arquivo, text="Selecionar Arquivo", style="Primary.TButton", command=selecionar_arquivo).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_arquivo, text="Carregar Colunas", style="Primary.TButton", command=carregar_colunas).pack(side=tk.LEFT, padx=5)

frame_pasta = ttk.Labelframe(frame_procv, text="Pasta de saída", style="Frame.TLabelframe", padding=10)
frame_pasta.pack(padx=12, pady=6, fill="x")

entry_pasta = tk.Entry(frame_pasta, textvariable=pasta_saida, font=fonte_entry, width=65,
                       bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, bd=1, relief="solid", highlightthickness=0)
entry_pasta.pack(side=tk.LEFT, padx=5, pady=3)

ttk.Button(frame_pasta, text="Selecionar Pasta", style="Warn.TButton", command=selecionar_pasta_saida).pack(side=tk.LEFT, padx=5)

frame_colunas = ttk.Labelframe(frame_procv, text="Colunas para comparação", style="Frame.TLabelframe", padding=10)
frame_colunas.pack(padx=12, pady=6, fill="x")

tk.Label(frame_colunas, text="Coluna A:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=0, column=0, padx=5, pady=3, sticky="w")
combo_colA = ttk.Combobox(frame_colunas, width=30, state="readonly")
combo_colA.grid(row=0, column=1, padx=5, pady=3)

tk.Label(frame_colunas, text="Coluna B:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=1, column=0, padx=5, pady=3, sticky="w")
combo_colB = ttk.Combobox(frame_colunas, width=30, state="readonly")
combo_colB.grid(row=1, column=1, padx=5, pady=3)

tk.Label(frame_procv, text="Tipo de comparação:", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_label).pack(pady=(8, 3))
combo_opcao = ttk.Combobox(frame_procv, width=40,
                           values=["O que tem na A e não tem na B", "O que tem na B e não tem na A"],
                           state="readonly")
combo_opcao.pack(pady=(0, 8))

ttk.Button(frame_procv, text="Executar Comparação", style="Accent.TButton", command=executar_comparacao).pack(pady=12)

progress = ttk.Progressbar(frame_procv, length=720, mode="determinate", style="Custom.Horizontal.TProgressbar")
progress.pack(pady=8)

frame_relatorio = ttk.Labelframe(frame_procv, text="Relatório", style="Frame.TLabelframe", padding=10)
frame_relatorio.pack(padx=12, pady=8, fill="both", expand=True)

txt_relatorio = tk.Text(frame_relatorio, width=95, height=15, font=("Consolas", 10),
                        bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                        bd=1, relief="solid", highlightthickness=1,
                        highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_relatorio.pack(fill="both", expand=True)

ttk.Button(frame_procv, text="Abrir pasta do arquivo gerado", style="Primary.TButton", command=abrir_pasta).pack(pady=(4, 12))


# =======================================================================
#                       ABA LIMPEZA DE DADOS (COM SCROLL)
# =======================================================================

base_empresas_path = tk.StringVar()
blocklist_c6_path = tk.StringVar()
nao_perturbe_1_path = tk.StringVar()
nao_perturbe_2_path = tk.StringVar()
nao_perturbe_3_path = tk.StringVar()
nao_perturbe_4_path = tk.StringVar()
out_dir_limpeza = tk.StringVar()
clean_mode_var = tk.StringVar(value="Simples")

limpeza_col_razao = tk.StringVar()
limpeza_col_tel = tk.StringVar()
limpeza_col_email = tk.StringVar()
limpeza_col_cnpj = tk.StringVar()
# Opções de telefone (pré-processamento na limpeza)
# - tel_has55_var: se os números já vêm com prefixo 55 (remover antes da limpeza)
# - add9_var: adiciona dígito 9 após o DDD (DD+8 -> DD9+8) quando aplicável
# - add55_var: adiciona prefixo 55 no início (código do país)
tel_has55_var = tk.StringVar(value="Não")
add9_var = tk.BooleanVar(value=False)
add55_var = tk.BooleanVar(value=False)


lbl_limpeza_title = tk.Label(frame_limpeza, text="Limpeza de dados", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_limpeza_title.pack(pady=(10, 2))

# --- (NOVO) Canvas com scrollbar vertical ---
limpeza_canvas = tk.Canvas(frame_limpeza, bg=BG_PRINCIPAL, highlightthickness=0)
limpeza_scrollbar = ttk.Scrollbar(frame_limpeza, orient="vertical", command=limpeza_canvas.yview)
limpeza_canvas.configure(yscrollcommand=limpeza_scrollbar.set)

limpeza_scrollbar.pack(side=tk.RIGHT, fill="y")
limpeza_canvas.pack(side=tk.LEFT, fill="both", expand=True)

frame_limpeza_scroll_content = tk.Frame(limpeza_canvas, bg=BG_PRINCIPAL)
limpeza_canvas_window = limpeza_canvas.create_window((0, 0), window=frame_limpeza_scroll_content, anchor="nw")

def _on_limpeza_configure(event):
    limpeza_canvas.configure(scrollregion=limpeza_canvas.bbox("all"))

def _on_limpeza_canvas_configure(event):
    # mantem o frame interno na largura do canvas
    try:
        limpeza_canvas.itemconfig(limpeza_canvas_window, width=event.width)
    except:
        pass

frame_limpeza_scroll_content.bind("<Configure>", _on_limpeza_configure)
limpeza_canvas.bind("<Configure>", _on_limpeza_canvas_configure)

# scroll com roda do mouse
def _on_mousewheel(event):
    # Windows
    if os.name == "nt":
        limpeza_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    else:
        # Linux/Mac (event.delta pode variar)
        limpeza_canvas.yview_scroll(int(-1 * (event.delta)), "units")

limpeza_canvas.bind_all("<MouseWheel>", _on_mousewheel)

# --- Layout original dentro do conteúdo rolável ---
frame_limpeza_main = tk.Frame(frame_limpeza_scroll_content, bg=BG_PRINCIPAL)
frame_limpeza_main.pack(fill="both", expand=True, padx=8, pady=8)

frame_limpeza_left = tk.Frame(frame_limpeza_main, bg=BG_PRINCIPAL)
frame_limpeza_left.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 6))

frame_limpeza_right = tk.Frame(frame_limpeza_main, bg=BG_PRINCIPAL)
frame_limpeza_right.pack(side=tk.LEFT, fill="both", expand=True, padx=(6, 0))

ttk.Button(frame_limpeza_left, text="Iniciar automação de limpeza", style="Accent.TButton", command=executar_limpeza_dados).pack(pady=(0, 10), anchor="w")

frame_modo = ttk.Labelframe(frame_limpeza_left, text="Modo de limpeza da Razão Social", style="Frame.TLabelframe", padding=10)
frame_modo.pack(padx=0, pady=6, fill="x")

tk.Label(frame_modo, text="Selecione o modo:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).pack(side=tk.LEFT, padx=(5, 5))
combo_modo = ttk.Combobox(frame_modo, textvariable=clean_mode_var, state="readonly", width=20,
                          values=["Simples", "Lemit", "Callix", "Tallos"])
combo_modo.pack(side=tk.LEFT, padx=5)
combo_modo.current(0)

# (NOVO) Opções de telefone (55 / dígito 9)
frame_tel_opts = ttk.Labelframe(frame_limpeza_left, text="Opções de telefone", style="Frame.TLabelframe", padding=10)
frame_tel_opts.pack(padx=0, pady=6, fill="x")

tk.Label(frame_tel_opts, text="Números têm 55?", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=0, column=0, padx=5, pady=5, sticky="w")
combo_has55 = ttk.Combobox(frame_tel_opts, textvariable=tel_has55_var, state="readonly", width=10, values=["Sim", "Não"])
combo_has55.grid(row=0, column=1, padx=5, pady=5, sticky="w")
combo_has55.current(1)

chk_add9 = ttk.Checkbutton(frame_tel_opts, text="Adicionar dígito 9 (DD+8 → DD9+8)", variable=add9_var)
chk_add9.grid(row=1, column=0, columnspan=2, padx=5, pady=4, sticky="w")

chk_add55 = ttk.Checkbutton(frame_tel_opts, text="Adicionar 55 na frente (código país)", variable=add55_var)
chk_add55.grid(row=2, column=0, columnspan=2, padx=5, pady=4, sticky="w")

# Arquivo base
frame_base = ttk.Labelframe(frame_limpeza_left, text='1) Arquivo "empresas bruto"', style="Frame.TLabelframe", padding=10)
frame_base.pack(padx=0, pady=6, fill="x")

entry_base = tk.Entry(frame_base, textvariable=base_empresas_path, font=fonte_entry, width=50,
                      bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, bd=1, relief="solid", highlightthickness=0)
entry_base.pack(side=tk.LEFT, padx=5, pady=3)

def selecionar_base_empresas():
    path = filedialog.askopenfilename(
        title='Selecione o arquivo "empresas bruto"',
        filetypes=[("Excel/CSV/TXT", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if path:
        base_empresas_path.set(path)

ttk.Button(frame_base, text="Selecionar", style="Primary.TButton", command=selecionar_base_empresas).pack(side=tk.LEFT, padx=5)

# Colunas
frame_cols_limpeza = ttk.Labelframe(frame_limpeza_left, text="2) Seleção de colunas (use escanear)", style="Frame.TLabelframe", padding=10)
frame_cols_limpeza.pack(padx=0, pady=6, fill="x")

ttk.Button(frame_cols_limpeza, text="Escanear colunas", style="Primary.TButton", command=escanear_colunas_limpeza).pack(anchor="w", pady=(0, 8))

tk.Label(frame_cols_limpeza, text="Razão Social:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
combo_limpeza_razao = ttk.Combobox(frame_cols_limpeza, textvariable=limpeza_col_razao, state="readonly")
combo_limpeza_razao.pack(fill="x", pady=2)

tk.Label(frame_cols_limpeza, text="Telefones:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
combo_limpeza_tel = ttk.Combobox(frame_cols_limpeza, textvariable=limpeza_col_tel, state="readonly")
combo_limpeza_tel.pack(fill="x", pady=2)

tk.Label(frame_cols_limpeza, text="E-mail:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
combo_limpeza_email = ttk.Combobox(frame_cols_limpeza, textvariable=limpeza_col_email, state="readonly")
combo_limpeza_email.pack(fill="x", pady=2)

tk.Label(frame_cols_limpeza, text="CNPJ:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
combo_limpeza_cnpj = ttk.Combobox(frame_cols_limpeza, textvariable=limpeza_col_cnpj, state="readonly")
combo_limpeza_cnpj.pack(fill="x", pady=2)

def _selecionar_arquivo_em_var(var: tk.StringVar, titulo: str):
    path = filedialog.askopenfilename(
        title=titulo,
        filetypes=[("Excel/CSV/TXT", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if path:
        var.set(path)

# Filtros
frame_f1 = ttk.Labelframe(frame_limpeza_left, text="3) Blocklist C6 (opcional)", style="Frame.TLabelframe", padding=10)
frame_f1.pack(padx=0, pady=6, fill="x")
tk.Entry(frame_f1, textvariable=blocklist_c6_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=50).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_f1, text="Selecionar", style="Primary.TButton", command=lambda: _selecionar_arquivo_em_var(blocklist_c6_path, "Selecione Blocklist C6")).pack(side=tk.LEFT, padx=5)

frame_f2 = ttk.Labelframe(frame_limpeza_left, text="4) Não Perturbe 1 (opcional)", style="Frame.TLabelframe", padding=10)
frame_f2.pack(padx=0, pady=6, fill="x")
tk.Entry(frame_f2, textvariable=nao_perturbe_1_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=50).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_f2, text="Selecionar", style="Primary.TButton", command=lambda: _selecionar_arquivo_em_var(nao_perturbe_1_path, "Selecione Não Perturbe 1")).pack(side=tk.LEFT, padx=5)

frame_f3 = ttk.Labelframe(frame_limpeza_left, text="5) Não Perturbe 2 (opcional)", style="Frame.TLabelframe", padding=10)
frame_f3.pack(padx=0, pady=6, fill="x")
tk.Entry(frame_f3, textvariable=nao_perturbe_2_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=50).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_f3, text="Selecionar", style="Primary.TButton", command=lambda: _selecionar_arquivo_em_var(nao_perturbe_2_path, "Selecione Não Perturbe 2")).pack(side=tk.LEFT, padx=5)

frame_f4 = ttk.Labelframe(frame_limpeza_left, text="6) Não Perturbe 3 (opcional)", style="Frame.TLabelframe", padding=10)
frame_f4.pack(padx=0, pady=6, fill="x")
tk.Entry(frame_f4, textvariable=nao_perturbe_3_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=50).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_f4, text="Selecionar", style="Primary.TButton", command=lambda: _selecionar_arquivo_em_var(nao_perturbe_3_path, "Selecione Não Perturbe 3")).pack(side=tk.LEFT, padx=5)

frame_f5 = ttk.Labelframe(frame_limpeza_left, text="7) Não Perturbe 4 (opcional)", style="Frame.TLabelframe", padding=10)
frame_f5.pack(padx=0, pady=6, fill="x")
tk.Entry(frame_f5, textvariable=nao_perturbe_4_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=50).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_f5, text="Selecionar", style="Primary.TButton", command=lambda: _selecionar_arquivo_em_var(nao_perturbe_4_path, "Selecione Não Perturbe 4")).pack(side=tk.LEFT, padx=5)

# Pasta saída
frame_out = ttk.Labelframe(frame_limpeza_left, text='8) Diretório de saída (onde salvar os 2 arquivos)', style="Frame.TLabelframe", padding=10)
frame_out.pack(padx=0, pady=6, fill="x")

entry_out = tk.Entry(frame_out, textvariable=out_dir_limpeza, font=fonte_entry, width=50,
                     bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, bd=1, relief="solid", highlightthickness=0)
entry_out.pack(side=tk.LEFT, padx=5, pady=3)

def selecionar_out_dir():
    path = filedialog.askdirectory(title='Selecione a pasta para salvar os arquivos finais')
    if path:
        out_dir_limpeza.set(path)

ttk.Button(frame_out, text="Selecionar Pasta", style="Warn.TButton", command=selecionar_out_dir).pack(side=tk.LEFT, padx=5)

frame_btns_limpeza = tk.Frame(frame_limpeza_left, bg=BG_PRINCIPAL)
frame_btns_limpeza.pack(pady=(10, 0), anchor="w")
ttk.Button(frame_btns_limpeza, text="Abrir pasta de saída", style="Primary.TButton", command=abrir_pasta_limpeza).pack(side=tk.LEFT, padx=6)

progress_limpeza = ttk.Progressbar(frame_limpeza_right, length=400, mode="determinate", style="Custom.Horizontal.TProgressbar")
progress_limpeza.pack(pady=(0, 8), padx=4, fill="x")

# (NOVO) área de gráficos
frame_graficos_limpeza = ttk.Labelframe(frame_limpeza_right, text="Gráficos (após execução)", style="Frame.TLabelframe", padding=10)
frame_graficos_limpeza.pack(padx=4, pady=4, fill="both", expand=False)
lbl_graf_info = tk.Label(frame_graficos_limpeza, text="Execute a limpeza para gerar os gráficos.",
                         bg=BG_FRAME, fg=FG_SECUNDARIO, font=("Segoe UI", 10))
lbl_graf_info.pack(fill="x")

frame_log_limpeza = ttk.Labelframe(frame_limpeza_right, text="Log / Relatório (tempo real)", style="Frame.TLabelframe", padding=10)
frame_log_limpeza.pack(padx=4, pady=4, fill="both", expand=True)

txt_log_limpeza = tk.Text(frame_log_limpeza, width=60, height=25, font=("Consolas", 10),
                          bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                          bd=1, relief="solid", highlightthickness=1,
                          highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_limpeza.pack(fill="both", expand=True)



# =======================================================================
#                       ABA LIMPEZA WHATSAPP (NOVA)
# =======================================================================

# Observação importante:
# - Esta aba NÃO usa automação de WhatsApp Web (Selenium/Playwright), pois isso é frágil e pode violar termos.
# - Aqui fazemos: normalização + padronização (+55 e dígito 9 quando necessário) + validação (phonenumbers) + filtro de móvel.
# - Se quiser checagem de "tem WhatsApp", o caminho seguro é via API oficial do WhatsApp Business Platform/Provedor (ex.: 360dialog),
#   o que exige credenciais e endpoints específicos.

wpp_base_path = tk.StringVar()
wpp_out_dir = tk.StringVar()
wpp_col_tel = tk.StringVar()

wpp_has9_var = tk.StringVar(value="Sim")     # Pergunta: número já tem 9?
wpp_has55_var = tk.StringVar(value="Não")    # Pergunta: número já tem 55?

# UI
lbl_wpp_title = tk.Label(frame_wpp, text="Limpeza WhatsApp (normalização + validação)", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_wpp_title.pack(pady=(10, 2))

lbl_wpp_sub = tk.Label(
    frame_wpp,
    text=(
        "Fluxo:\n"
        "1) Você seleciona a planilha e escolhe a coluna de telefone\n"
        "2) Responde se os números já possuem 55 e/ou dígito 9\n"
        "3) O sistema normaliza, valida (phonenumbers) e filtra somente números móveis\n"
        "4) Gera 2 arquivos: whatsapp_validos.xlsx e whatsapp_excluidos.xlsx (com motivo)\n"
    ),
    bg=BG_PRINCIPAL, fg=FG_SECUNDARIO, font=("Segoe UI", 10), justify="center"
)
lbl_wpp_sub.pack(pady=(0, 10))

frame_wpp_main = tk.Frame(frame_wpp, bg=BG_PRINCIPAL)
frame_wpp_main.pack(fill="both", expand=True, padx=8, pady=8)

frame_wpp_left = tk.Frame(frame_wpp_main, bg=BG_PRINCIPAL)
frame_wpp_left.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 6))

frame_wpp_right = tk.Frame(frame_wpp_main, bg=BG_PRINCIPAL)
frame_wpp_right.pack(side=tk.LEFT, fill="both", expand=True, padx=(6, 0))

def log_wpp(msg: str):
    txt_log_wpp.insert(tk.END, msg + "\n")
    txt_log_wpp.see(tk.END)
    janela.update_idletasks()

def selecionar_base_wpp():
    path = filedialog.askopenfilename(
        title='Selecione a planilha para limpeza de WhatsApp',
        filetypes=[("Excel/CSV/TXT", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if path:
        wpp_base_path.set(path)

def selecionar_out_dir_wpp():
    path = filedialog.askdirectory(title='Selecione a pasta para salvar os arquivos do WhatsApp')
    if path:
        wpp_out_dir.set(path)

def escanear_colunas_wpp():
    try:
        in_path = wpp_base_path.get().strip()
        if not in_path:
            messagebox.showwarning("Aviso", "Selecione a planilha primeiro.")
            return
        log_wpp("🔎 Escaneando colunas...")
        df = read_table(in_path)
        cols = list(df.columns)
        combo_wpp_tel["values"] = cols

        # sugestão
        normals = {normalize_col_name(c): c for c in cols}
        sug = ""
        for key in ["telefone", "telefones", "tel", "fone", "celular", "whatsapp", "wpp"]:
            kk = normalize_col_name(key)
            if kk in normals:
                sug = normals[kk]
                break
        if not sug:
            for c in cols:
                if any(k in normalize_col_name(c) for k in ["tel", "fone", "cel", "wpp", "whats"]):
                    sug = c
                    break
        if sug:
            wpp_col_tel.set(sug)

        log_wpp(f"✅ Colunas carregadas: {len(cols)}")
        messagebox.showinfo("OK", "Colunas carregadas. Selecione a coluna de telefone.")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao escanear colunas.\n\n{e}")

def abrir_pasta_wpp():
    pasta = wpp_out_dir.get().strip() or os.path.dirname(wpp_base_path.get().strip() or "")
    if not pasta:
        messagebox.showwarning("Aviso", "Nenhuma pasta disponível.")
        return
    try:
        if sys.platform == "darwin":
            subprocess.Popen(["open", pasta])
        elif os.name == "nt":
            subprocess.Popen(f'explorer "{pasta}"')
        else:
            subprocess.Popen(["xdg-open", pasta])
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível abrir a pasta.\n\n{e}")

# --- Helpers WhatsApp ---
try:
    import phonenumbers
    from phonenumbers import PhoneNumberType
except Exception:
    phonenumbers = None
    PhoneNumberType = None

def _digits_only(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))

def _apply_has55_rule(digits: str, has55: str) -> str:
    # Se usuário disse "Sim", removemos 55 somente se estiver no começo.
    if has55 == "Sim" and digits.startswith("55"):
        return digits[2:]
    return digits

def _ensure_add55(digits_local: str) -> str:
    # Adiciona 55 se estiver ausente e parecer BR (10 ou 11 dígitos)
    if digits_local.startswith("55"):
        return digits_local
    if len(digits_local) in (10, 11):
        return "55" + digits_local
    return digits_local

def _ensure_add9_local(digits_local: str, has9: str) -> str:
    # has9 == "Não" => se for 10 dígitos (DD + 8), vira 11 (DD + 9 + 8)
    if has9 == "Não" and len(digits_local) == 10:
        ddd = digits_local[:2]
        rest = digits_local[2:]
        return ddd + "9" + rest
    return digits_local

def _format_e164(digits: str) -> str:
    # Retorna no formato +5511999999999 quando possível
    digits = _digits_only(digits)
    if digits.startswith("55") and len(digits) in (12, 13):
        return "+" + digits
    if len(digits) in (10, 11):
        return "+55" + digits
    if digits.startswith("+"):
        return digits
    return "+" + digits if digits else ""

def _phonenumbers_validate_br(digits_e164: str):
    """
    Retorna: (is_valid, is_mobile, tipo_str, motivo)
    """
    if not phonenumbers:
        return (False, False, "Indisponível", "Biblioteca phonenumbers não instalada")
    if not digits_e164:
        return (False, False, "Inválido", "Telefone vazio")
    try:
        # parse aceita +E164
        p = phonenumbers.parse(digits_e164, None)
        if not phonenumbers.is_valid_number(p):
            return (False, False, "Inválido", "Número inválido (phonenumbers)")
        t = phonenumbers.number_type(p)
        # Para evitar falsos negativos: BR às vezes vem como FIXED_LINE_OR_MOBILE
        is_mobile = t in (PhoneNumberType.MOBILE, PhoneNumberType.FIXED_LINE_OR_MOBILE)
        tipo_str = "Móvel" if is_mobile else "Não móvel"
        return (True, is_mobile, tipo_str, "" if is_mobile else "Não é móvel")
    except Exception as e:
        return (False, False, "Inválido", f"Erro ao validar: {e}")

def executar_limpeza_wpp():
    try:
        txt_log_wpp.delete("1.0", tk.END)
        progress_wpp["value"] = 0
        janela.update_idletasks()

        in_path = wpp_base_path.get().strip()
        if not in_path:
            messagebox.showwarning("Aviso", "Selecione a planilha.")
            return

        col_tel = wpp_col_tel.get().strip()
        if not col_tel:
            messagebox.showwarning("Aviso", "Selecione a coluna de telefone (use Escanear colunas).")
            return

        out_dir = wpp_out_dir.get().strip() or os.path.dirname(in_path)
        os.makedirs(out_dir, exist_ok=True)

        has55 = wpp_has55_var.get()
        has9 = wpp_has9_var.get()

        log_wpp("=== Limpeza WhatsApp ===")
        log_wpp(f"📄 Arquivo: {in_path}")
        log_wpp(f"📌 Coluna telefone: {col_tel}")
        log_wpp(f"❓ Números têm 55? {has55} (Se Sim: remove 55 no início)")
        log_wpp(f"❓ Números têm dígito 9? {has9} (Se Não: adiciona 9 em DD+8)")
        if not phonenumbers:
            log_wpp("⚠️ Biblioteca 'phonenumbers' não está instalada. A validação ficará indisponível.")
            log_wpp("   Instale com: pip install phonenumbers\n")

        progress_wpp["value"] = 10
        janela.update_idletasks()

        df_raw = read_table(in_path)
        if col_tel not in df_raw.columns:
            messagebox.showerror("Erro", f"Coluna '{col_tel}' não existe no arquivo.")
            return

        progress_wpp["value"] = 25
        janela.update_idletasks()

        # Processa
        out_rows = []
        total = len(df_raw)
        for i, raw_tel in enumerate(df_raw[col_tel].astype(str).tolist(), start=1):
            digits = _digits_only(raw_tel)
            motivo = ""

            if not digits:
                out_rows.append({
                    "Telefone_original": raw_tel,
                    "Telefone_normalizado": "",
                    "Telefone_E164": "",
                    "Valido": "Não",
                    "Tipo": "Inválido",
                    "Motivo": "Telefone vazio"
                })
                continue

            # 1) remove 55 se usuário disse que já tem
            local = _apply_has55_rule(digits, has55)

            # 2) adiciona 9 se necessário
            local = _ensure_add9_local(local, has9)

            # 3) garante 55 (para E.164) — aqui sempre padronizamos em +55...
            full = _ensure_add55(local)
            e164 = _format_e164(full)

            # 4) valida e filtra móvel
            is_valid, is_mobile, tipo_str, motivo_v = _phonenumbers_validate_br(e164)
            if not is_valid:
                motivo = motivo_v or "Inválido"
            elif not is_mobile:
                motivo = motivo_v or "Não é móvel"

            out_rows.append({
                "Telefone_original": raw_tel,
                "Telefone_normalizado": full,
                "Telefone_E164": e164,
                "Valido": "Sim" if (is_valid and is_mobile) else "Não",
                "Tipo": tipo_str,
                "Motivo": motivo
            })

            if i % 500 == 0:
                progress_wpp["value"] = 25 + int((i / max(total, 1)) * 55)
                janela.update_idletasks()

        df_out = pd.DataFrame(out_rows)

        # separa
        df_validos = df_out[df_out["Valido"] == "Sim"].copy()
        df_excluidos = df_out[df_out["Valido"] != "Sim"].copy()

        progress_wpp["value"] = 85
        janela.update_idletasks()

        out_valid = os.path.join(out_dir, "whatsapp_validos.xlsx")
        out_excl = os.path.join(out_dir, "whatsapp_excluidos.xlsx")
        safe_remove_file(out_valid)
        safe_remove_file(out_excl)

        save_to_excel(df_validos, out_valid)
        save_to_excel(df_excluidos, out_excl)

        progress_wpp["value"] = 100
        janela.update_idletasks()

        log_wpp("🎉 Concluído!")
        log_wpp(f"✅ Válidos (móvel + válido): {len(df_validos)} → {out_valid}")
        log_wpp(f"✅ Excluídos: {len(df_excluidos)} → {out_excl}")

        messagebox.showinfo("Concluído", f"Limpeza WhatsApp finalizada!\n\nVálidos: {len(df_validos)}\nExcluídos: {len(df_excluidos)}\n\nSaída: {out_dir}")

    except Exception as e:
        log_wpp(f"\n❌ Erro fatal: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro.\n\n{e}")

# --- Widgets da aba ---
frame_wpp_file = ttk.Labelframe(frame_wpp_left, text="1) Planilha", style="Frame.TLabelframe", padding=10)
frame_wpp_file.pack(fill="x", pady=6)

tk.Entry(frame_wpp_file, textvariable=wpp_base_path, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=55).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_wpp_file, text="Selecionar", style="Primary.TButton", command=selecionar_base_wpp).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_wpp_file, text="Escanear colunas", style="Primary.TButton", command=escanear_colunas_wpp).pack(side=tk.LEFT, padx=5)

frame_wpp_cols = ttk.Labelframe(frame_wpp_left, text="2) Coluna de telefone", style="Frame.TLabelframe", padding=10)
frame_wpp_cols.pack(fill="x", pady=6)

tk.Label(frame_wpp_cols, text="Telefone:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
combo_wpp_tel = ttk.Combobox(frame_wpp_cols, textvariable=wpp_col_tel, state="readonly")
combo_wpp_tel.pack(fill="x", pady=2)

frame_wpp_opts = ttk.Labelframe(frame_wpp_left, text="3) Perguntas para padronização", style="Frame.TLabelframe", padding=10)
frame_wpp_opts.pack(fill="x", pady=6)

row1 = tk.Frame(frame_wpp_opts, bg=BG_FRAME)
row1.pack(fill="x", pady=2)
tk.Label(row1, text="Número já tem dígito 9?", bg=BG_FRAME, fg=FG_TEXTO).pack(side=tk.LEFT, padx=5)
ttk.Combobox(row1, textvariable=wpp_has9_var, state="readonly", width=10, values=["Sim", "Não"]).pack(side=tk.LEFT, padx=5)

row2 = tk.Frame(frame_wpp_opts, bg=BG_FRAME)
row2.pack(fill="x", pady=2)
tk.Label(row2, text="Número já tem 55?", bg=BG_FRAME, fg=FG_TEXTO).pack(side=tk.LEFT, padx=5)
ttk.Combobox(row2, textvariable=wpp_has55_var, state="readonly", width=10, values=["Sim", "Não"]).pack(side=tk.LEFT, padx=5)

frame_wpp_out = ttk.Labelframe(frame_wpp_left, text="4) Pasta de saída", style="Frame.TLabelframe", padding=10)
frame_wpp_out.pack(fill="x", pady=6)

tk.Entry(frame_wpp_out, textvariable=wpp_out_dir, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=55).pack(side=tk.LEFT, padx=5, pady=3)
ttk.Button(frame_wpp_out, text="Selecionar pasta", style="Warn.TButton", command=selecionar_out_dir_wpp).pack(side=tk.LEFT, padx=5)

ttk.Button(frame_wpp_left, text="Executar limpeza WhatsApp", style="Accent.TButton", command=executar_limpeza_wpp).pack(pady=(10, 5), anchor="w")
ttk.Button(frame_wpp_left, text="Abrir pasta de saída", style="Primary.TButton", command=abrir_pasta_wpp).pack(pady=(0, 5), anchor="w")

progress_wpp = ttk.Progressbar(frame_wpp_right, length=400, mode="determinate", style="Custom.Horizontal.TProgressbar")
progress_wpp.pack(pady=(0, 8), padx=4, fill="x")

frame_log_wpp = ttk.Labelframe(frame_wpp_right, text="Log WhatsApp", style="Frame.TLabelframe", padding=10)
frame_log_wpp.pack(padx=4, pady=4, fill="both", expand=True)

txt_log_wpp = tk.Text(frame_log_wpp, width=60, height=25, font=("Consolas", 10),
                      bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                      bd=1, relief="solid", highlightthickness=1,
                      highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_wpp.pack(fill="both", expand=True)



# =======================================================================
#                       ABA ROBÔ C6
# =======================================================================

robo_arquivos: List[str] = []
robo_bat_path = tk.StringVar()
robo_resultado_dir = tk.StringVar()
robo_modo_var = tk.StringVar(value="Simples")

lbl_robo_title = tk.Label(frame_robo, text="Robô C6", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_robo_title.pack(pady=(10, 2))

lbl_robo_sub = tk.Label(
    frame_robo,
    text=(
        "Automatiza a execução de um .BAT do C6 para múltiplas planilhas.\n"
        "• Lê várias planilhas (até 20 mil linhas cada)\n"
        "• Envia uma por vez para a pasta do .BAT\n"
        "• Executa o .BAT, envia ENTER ao final e espera 6 minutos\n"
        "• Consolida resultados, filtra somente 'Novo cliente' (removendo 'Nao disponivel')\n"
        "• Gera saída em modo Lemit (1 arquivo) ou Simples (arquivos de 5.000 linhas)"
    ),
    bg=BG_PRINCIPAL,
    fg=FG_SECUNDARIO,
    font=("Segoe UI", 10),
    justify="center"
)
lbl_robo_sub.pack(pady=(0, 10))

frame_robo_main = tk.Frame(frame_robo, bg=BG_PRINCIPAL)
frame_robo_main.pack(fill="both", expand=True, padx=8, pady=8)

frame_robo_left = tk.Frame(frame_robo_main, bg=BG_PRINCIPAL)
frame_robo_left.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 6))

frame_robo_right = tk.Frame(frame_robo_main, bg=BG_PRINCIPAL)
frame_robo_right.pack(side=tk.LEFT, fill="both", expand=True, padx=(6, 0))

frame_robo_files = ttk.Labelframe(frame_robo_left, text="1) Planilhas de entrada (até 20 mil linhas cada)", style="Frame.TLabelframe", padding=10)
frame_robo_files.pack(fill="x", pady=6)

lbl_arquivos_robo = tk.Label(frame_robo_files, text="Nenhum arquivo selecionado.", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label, justify="left")
lbl_arquivos_robo.pack(anchor="w", pady=(0, 4))

ttk.Button(frame_robo_files, text="Selecionar planilhas", style="Primary.TButton", command=selecionar_arquivos_robo).pack(anchor="w", pady=4)

frame_robo_bat = ttk.Labelframe(frame_robo_left, text="2) Arquivo .BAT do C6", style="Frame.TLabelframe", padding=10)
frame_robo_bat.pack(fill="x", pady=6)

lbl_bat_robo = tk.Label(frame_robo_bat, text="Nenhum .BAT selecionado.", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label, justify="left")
lbl_bat_robo.pack(anchor="w", pady=(0, 4))

ttk.Button(frame_robo_bat, text="Selecionar .BAT", style="Warn.TButton", command=selecionar_bat_robo).pack(anchor="w", pady=4)

frame_robo_resultado = ttk.Labelframe(frame_robo_left, text="3) Pasta de resultados do .BAT", style="Frame.TLabelframe", padding=10)
frame_robo_resultado.pack(fill="x", pady=6)

lbl_pasta_resultado = tk.Label(
    frame_robo_resultado,
    text="Nenhuma pasta de resultados selecionada.\n(se não selecionar, tente usar 'pasta_do_bat/resultado')",
    bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label, wraplength=350, justify="left"
)
lbl_pasta_resultado.pack(pady=(0, 5), anchor="w")

ttk.Button(frame_robo_resultado, text="Selecionar pasta resultado", style="Primary.TButton", command=selecionar_resultado_robo).pack(pady=5, anchor="w")

frame_robo_modo = ttk.Labelframe(frame_robo_left, text="4) Modo de tratamento final (Lemit / Simples)", style="Frame.TLabelframe", padding=10)
frame_robo_modo.pack(fill="x", pady=6)

tk.Label(frame_robo_modo, text="Arquivo é para:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).pack(side=tk.LEFT, padx=(5, 5))
combo_robo_modo = ttk.Combobox(frame_robo_modo, textvariable=robo_modo_var, state="readonly", width=20, values=["Lemit", "Simples"])
combo_robo_modo.pack(side=tk.LEFT, padx=5)
combo_robo_modo.current(1)

ttk.Button(frame_robo_left, text="Executar Robô C6", style="Accent.TButton", command=executar_robo_c6).pack(pady=(10, 5), anchor="w")

progress_robo = ttk.Progressbar(frame_robo_right, length=400, mode="determinate", style="Custom.Horizontal.TProgressbar")
progress_robo.pack(pady=(0, 8), padx=4, fill="x")

frame_log_robo = ttk.Labelframe(frame_robo_right, text="Log do Robô C6", style="Frame.TLabelframe", padding=10)
frame_log_robo.pack(padx=4, pady=4, fill="both", expand=True)

txt_log_robo = tk.Text(frame_log_robo, width=60, height=25, font=("Consolas", 10),
                       bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                       bd=1, relief="solid", highlightthickness=1,
                       highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_robo.pack(fill="both", expand=True)


# =======================================================================
#                       ABA MANIPULAÇÃO
# =======================================================================

manip_arquivos: List[str] = []
manip_modo_var = tk.StringVar(value="juntar")  # juntar | separar
manip_out_dir = tk.StringVar()
manip_linhas_por_planilha = tk.StringVar(value="5000")

lbl_manip_title = tk.Label(frame_manip, text="Manipulação de planilhas", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_manip_title.pack(pady=(10, 2))

frame_manip_top = tk.Frame(frame_manip, bg=BG_PRINCIPAL)
frame_manip_top.pack(fill="both", expand=True, padx=12, pady=12)

frame_manip_left = tk.Frame(frame_manip_top, bg=BG_PRINCIPAL)
frame_manip_left.pack(side=tk.LEFT, fill="y", padx=(0, 10))

frame_manip_right = tk.Frame(frame_manip_top, bg=BG_PRINCIPAL)
frame_manip_right.pack(side=tk.LEFT, fill="both", expand=True)

frame_manip_mode = ttk.Labelframe(frame_manip_left, text="Modo", style="Frame.TLabelframe", padding=10)
frame_manip_mode.pack(fill="x", pady=6)

ttk.Radiobutton(frame_manip_mode, text="Juntar planilhas", variable=manip_modo_var, value="juntar").pack(anchor="w")
ttk.Radiobutton(frame_manip_mode, text="Separar em várias planilhas", variable=manip_modo_var, value="separar").pack(anchor="w")

frame_manip_chunk = ttk.Labelframe(frame_manip_left, text="Separar", style="Frame.TLabelframe", padding=10)
frame_manip_chunk.pack(fill="x", pady=6)

tk.Label(frame_manip_chunk, text="Linhas por planilha:", bg=BG_FRAME, fg=FG_TEXTO).pack(anchor="w")
tk.Entry(frame_manip_chunk, textvariable=manip_linhas_por_planilha,
         bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=12).pack(anchor="w", pady=4)

frame_manip_files = ttk.Labelframe(frame_manip_left, text="Arquivos", style="Frame.TLabelframe", padding=10)
frame_manip_files.pack(fill="x", pady=6)

lbl_manip_arquivos = tk.Label(frame_manip_files, text="Nenhum arquivo selecionado.", bg=BG_FRAME, fg=FG_TEXTO)
lbl_manip_arquivos.pack(anchor="w", pady=(0, 5))

ttk.Button(frame_manip_files, text="Selecionar planilhas", style="Primary.TButton", command=selecionar_arquivos_manip).pack(anchor="w", pady=3)

frame_manip_out = ttk.Labelframe(frame_manip_left, text="Saída", style="Frame.TLabelframe", padding=10)
frame_manip_out.pack(fill="x", pady=6)

lbl_manip_saida = tk.Label(frame_manip_out, text="Saída: (não selecionada)", bg=BG_FRAME, fg=FG_TEXTO)
lbl_manip_saida.pack(anchor="w", pady=(0, 5))

ttk.Button(frame_manip_out, text="Selecionar pasta", style="Warn.TButton", command=selecionar_out_dir_manip).pack(anchor="w", pady=3)

ttk.Button(frame_manip_left, text="Executar", style="Accent.TButton", command=executar_manipulacao).pack(pady=(10, 0), anchor="w")

frame_log_manip = ttk.Labelframe(frame_manip_right, text="Log", style="Frame.TLabelframe", padding=10)
frame_log_manip.pack(fill="both", expand=True)

txt_log_manip = tk.Text(frame_log_manip, width=80, height=30, font=("Consolas", 10),
                        bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                        bd=1, relief="solid", highlightthickness=1,
                        highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_manip.pack(fill="both", expand=True)


# =======================================================================
#                      ABA BANCO DE DADOS (IMPORT + VIEW)
# =======================================================================

import_tabela_var = tk.StringVar(value="empresas")
import_arquivo_path = tk.StringVar()
view_tabela_var = tk.StringVar(value="empresas")

def log_bd(msg: str):
    txt_log_bd.insert(tk.END, msg + "\n")
    txt_log_bd.see(tk.END)
    janela.update_idletasks()

frame_bd_main = tk.Frame(frame_bd, bg=BG_PRINCIPAL)
frame_bd_main.pack(fill="both", expand=True, padx=8, pady=8)

frame_bd_left = tk.Frame(frame_bd_main, bg=BG_PRINCIPAL)
frame_bd_left.pack(side=tk.LEFT, fill="y", expand=False, padx=(0, 6))

frame_bd_right = tk.Frame(frame_bd_main, bg=BG_PRINCIPAL)
frame_bd_right.pack(side=tk.LEFT, fill="both", expand=True, padx=(6, 0))

lbl_bd_title = tk.Label(frame_bd_left, text="Importar dados para o banco", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_bd_title.pack(pady=(0, 10))

frame_import_cfg = ttk.Labelframe(frame_bd_left, text="Configuração de importação", style="Frame.TLabelframe", padding=10)
frame_import_cfg.pack(fill="x", pady=6)

tk.Label(frame_import_cfg, text="Tabela destino:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=0, column=0, padx=5, pady=5, sticky="w")

combo_tab_dest = ttk.Combobox(
    frame_import_cfg,
    textvariable=import_tabela_var,
    state="readonly",
    width=25,
    values=["empresas", "block_list_c6", "block_list_b2b", "nao_perturbe", "cnais_aceitos", "lemit_relatorio"]
)
combo_tab_dest.grid(row=0, column=1, padx=5, pady=5)

frame_import_file = ttk.Labelframe(frame_bd_left, text="Arquivo bruto (Excel/CSV/TXT)", style="Frame.TLabelframe", padding=10)
frame_import_file.pack(fill="x", pady=6)

entry_import_file = tk.Entry(frame_import_file, textvariable=import_arquivo_path, font=fonte_entry, width=40,
                             bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, bd=1, relief="solid", highlightthickness=0)
entry_import_file.pack(side=tk.LEFT, padx=5, pady=3)

def selecionar_arquivo_import():
    path = filedialog.askopenfilename(
        title='Selecione o arquivo bruto',
        filetypes=[("Excel/CSV/TXT", "*.xlsx *.xls *.csv *.txt"), ("Todos os arquivos", "*.*")]
    )
    if path:
        import_arquivo_path.set(path)

ttk.Button(frame_import_file, text="Selecionar", style="Primary.TButton", command=selecionar_arquivo_import).pack(side=tk.LEFT, padx=5)

def importar_arquivo_para_tabela():
    global db_engine, db_connected

    if not db_connected or db_engine is None:
        messagebox.showwarning("Aviso", "Banco de dados não está conectado. Vá na aba 'Conexão BD' e conecte primeiro.")
        return

    tabela = import_tabela_var.get()
    path = import_arquivo_path.get().strip()

    if not tabela:
        messagebox.showwarning("Aviso", "Selecione a tabela destino.")
        return
    if not path:
        messagebox.showwarning("Aviso", "Selecione um arquivo bruto para importar.")
        return

    try:
        log_bd(f"📂 Lendo arquivo bruto: {path}")
        df_raw = read_table(path)
        log_bd(f"✅ Arquivo lido com {len(df_raw)} linhas e {len(df_raw.columns)} colunas.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao ler arquivo.\n\n{e}")
        log_bd(f"❌ Erro ao ler arquivo: {e}")
        return

    normals = {normalize_col_name(c): c for c in df_raw.columns}

    try:
        if tabela == "empresas":
            log_bd("🧩 Preparando dados para tabela EMPRESAS...")

            cnpj_col = pick_col(normals, ["cnpj"])
            razao_col = pick_col(normals, ["razao social", "razao_social", "razao"])
            situacao_col = pick_col(normals, ["situacao cadastral", "situacao"])
            uf_col = pick_col(normals, ["uf", "estado"])
            data_abertura_col = pick_col(normals, ["data abertura", "dt abertura", "abertura"])
            telefones_col = pick_col(normals, ["telefones", "telefone"])
            tel1_col = pick_col(normals, ["telefone1", "tel1"])
            tel2_col = pick_col(normals, ["telefone2", "tel2"])
            email_col = pick_col(normals, ["email", "e-mail"])
            capital_col = pick_col(normals, ["capital social"])
            socios_col = pick_col(normals, ["socios", "sócios"])
            ultimo_uso_col = pick_col(normals, ["ultimo uso", "ultimo_uso", "ultimo contato"])
            plataforma_col = pick_col(normals, ["plataforma usada", "plataforma", "origem"])

            df_emp = pd.DataFrame()

            df_emp["cnpj"] = df_raw[cnpj_col].apply(normalize_cnpj) if cnpj_col else None
            df_emp["razao_social"] = df_raw[razao_col].astype(str) if razao_col else None
            df_emp["situacao_cadastral"] = df_raw[situacao_col].astype(str) if situacao_col else None
            df_emp["uf"] = df_raw[uf_col].astype(str).str[:2].str.upper() if uf_col else None

            if data_abertura_col:
                dt = pd.to_datetime(df_raw[data_abertura_col], errors="coerce", dayfirst=True)
                df_emp["data_abertura"] = dt.dt.date
            else:
                df_emp["data_abertura"] = None

            tel1 = None
            tel2 = None
            if telefones_col:
                t1, t2 = zip(*df_raw[telefones_col].map(split_telefones_field))
                tel1 = pd.Series(t1).apply(normalize_phone)
                tel2 = pd.Series(t2).apply(normalize_phone)
            if tel1_col and tel1 is None:
                tel1 = df_raw[tel1_col].apply(normalize_phone)
            if tel2_col and tel2 is None:
                tel2 = df_raw[tel2_col].apply(normalize_phone)

            df_emp["telefone1"] = tel1 if tel1 is not None else None
            df_emp["telefone2"] = tel2 if tel2 is not None else None
            df_emp["email"] = df_raw[email_col].astype(str) if email_col else None

            if capital_col:
                cap_series = df_raw[capital_col].astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
                df_emp["capital_social"] = pd.to_numeric(cap_series, errors="coerce")
            else:
                df_emp["capital_social"] = None

            df_emp["socios"] = df_raw[socios_col].astype(str) if socios_col else None

            if ultimo_uso_col:
                dt_ult = pd.to_datetime(df_raw[ultimo_uso_col], errors="coerce", dayfirst=True)
                df_emp["ultimo_uso"] = dt_ult
            else:
                df_emp["ultimo_uso"] = None

            df_emp["plataforma_usada"] = df_raw[plataforma_col].astype(str) if plataforma_col else None

            before = len(df_emp)
            df_emp = df_emp[df_emp["cnpj"].notna() & (df_emp["cnpj"] != "")]
            log_bd(f"➡️ Removidas {before - len(df_emp)} linhas sem CNPJ.")
            log_bd(f"📥 Preparado {len(df_emp)} registros para INSERT em 'empresas'.")

            df_emp.to_sql("empresas", db_engine, if_exists="append", index=False, chunksize=5000, method="multi")
            log_bd("✅ Importação para 'empresas' concluída.")

        elif tabela in ["block_list_c6", "block_list_b2b", "nao_perturbe"]:
            log_bd(f"🧩 Preparando dados para tabela {tabela.upper()}...")

            tel_col = None
            for c in df_raw.columns:
                if "tel" in normalize_col_name(c):
                    tel_col = c
                    break
            if not tel_col:
                raise ValueError("Não foi encontrada nenhuma coluna de telefone no arquivo.")

            df_tel = pd.DataFrame()
            df_tel["telefone"] = df_raw[tel_col].apply(normalize_phone)
            before = len(df_tel)
            df_tel = df_tel[df_tel["telefone"] != ""].drop_duplicates()
            log_bd(f"➡️ {before - len(df_tel)} linhas removidas (vazias/duplicadas).")

            df_tel.to_sql(tabela, db_engine, if_exists="append", index=False, chunksize=10000, method="multi")
            log_bd(f"✅ Importação para '{tabela}' concluída ({len(df_tel)} registros).")

        elif tabela == "cnais_aceitos":
            log_bd("🧩 Preparando dados para tabela CNAIS_ACEITOS...")

            cnai_col = None
            for c in df_raw.columns:
                norm = normalize_col_name(c)
                if any(k in norm for k in ["cnae", "cnai", "codigo"]):
                    cnai_col = c
                    break
            if not cnai_col:
                raise ValueError("Não foi encontrada nenhuma coluna de CNAI/CNAE no arquivo.")

            df_cnai = pd.DataFrame()
            df_cnai["cnai"] = df_raw[cnai_col].astype(str).str.strip()
            before = len(df_cnai)
            df_cnai = df_cnai[df_cnai["cnai"] != ""].drop_duplicates()
            log_bd(f"➡️ {before - len(df_cnai)} linhas removidas (vazias/duplicadas).")

            df_cnai.to_sql("cnais_aceitos", db_engine, if_exists="append", index=False, chunksize=10000, method="multi")
            log_bd(f"✅ Importação para 'cnais_aceitos' concluída ({len(df_cnai)} registros).")

        elif tabela == "lemit_relatorio":
            log_bd("🧩 Preparando dados para tabela LEMIT_RELATORIO...")

            contato_col = pick_col(normals, ["contato", "nome contato", "responsavel"])
            telefone_col = pick_col(normals, ["telefone", "tel", "telefone contato"])
            desc_col = pick_col(normals, ["descricao interacao", "descrição interacao", "descricao", "observacao"])
            cnpj_col = pick_col(normals, ["cnpj"])
            email_col = pick_col(normals, ["email", "e-mail"])
            data_abertura_col = pick_col(normals, ["data abertura", "dt abertura", "abertura"])
            uf_col = pick_col(normals, ["uf", "estado"])

            df_lr = pd.DataFrame()
            df_lr["contato"] = df_raw[contato_col].astype(str) if contato_col else None
            df_lr["telefone"] = df_raw[telefone_col].apply(normalize_phone) if telefone_col else None
            df_lr["descricao_interacao"] = df_raw[desc_col].astype(str) if desc_col else None
            df_lr["cnpj"] = df_raw[cnpj_col].apply(normalize_cnpj) if cnpj_col else None
            df_lr["email"] = df_raw[email_col].astype(str) if email_col else None

            if data_abertura_col:
                dt = pd.to_datetime(df_raw[data_abertura_col], errors="coerce", dayfirst=True)
                df_lr["data_abertura"] = dt.dt.date
            else:
                df_lr["data_abertura"] = None

            df_lr["uf"] = df_raw[uf_col].astype(str).str[:2].str.upper() if uf_col else None

            before = len(df_lr)
            df_lr = df_lr[df_lr["cnpj"].notna() | df_lr["telefone"].notna()]
            log_bd(f"➡️ {before - len(df_lr)} linhas removidas (sem CNPJ e sem telefone).")

            df_lr.to_sql("lemit_relatorio", db_engine, if_exists="append", index=False, chunksize=5000, method="multi")
            log_bd(f"✅ Importação para 'lemit_relatorio' concluída ({len(df_lr)} registros).")

        else:
            raise ValueError(f"Tabela '{tabela}' não tratada na importação.")

        messagebox.showinfo("Concluído", f"Importação para '{tabela}' finalizada com sucesso.")

    except Exception as e:
        log_bd(f"❌ Erro durante importação: {e}")
        messagebox.showerror("Erro", f"Erro durante importação.\n\n{e}")

ttk.Button(frame_bd_left, text="Importar arquivo para tabela", style="Accent.TButton", command=importar_arquivo_para_tabela).pack(pady=(10, 5))


# Visualização
lbl_view_title = tk.Label(frame_bd_right, text="Visualizar dados das tabelas", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_view_title.pack(pady=(0, 10))

frame_view_top = tk.Frame(frame_bd_right, bg=BG_PRINCIPAL)
frame_view_top.pack(fill="x", pady=(0, 5), padx=4)

tk.Label(frame_view_top, text="Tabela:", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_label).pack(side=tk.LEFT, padx=(0, 5))

combo_tab_view = ttk.Combobox(
    frame_view_top,
    textvariable=view_tabela_var,
    state="readonly",
    width=25,
    values=["empresas", "block_list_c6", "block_list_b2b", "nao_perturbe", "cnais_aceitos", "lemit_relatorio"]
)
combo_tab_view.pack(side=tk.LEFT, padx=(0, 5))
combo_tab_view.current(0)

def visualizar_tabela():
    global db_engine, db_connected

    if not db_connected or db_engine is None:
        messagebox.showwarning("Aviso", "Banco de dados não está conectado. Vá na aba 'Conexão BD' e conecte primeiro.")
        return

    tabela = view_tabela_var.get()
    if not tabela:
        messagebox.showwarning("Aviso", "Selecione uma tabela para visualizar.")
        return

    try:
        query = f"SELECT * FROM {tabela} LIMIT 200"
        df = pd.read_sql_query(query, db_engine)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao consultar tabela.\n\n{e}")
        return

    tree_bd.delete(*tree_bd.get_children())
    tree_bd["columns"] = list(df.columns)
    tree_bd["show"] = "headings"

    for col in df.columns:
        tree_bd.heading(col, text=col)
        tree_bd.column(col, width=120, anchor="w")

    for _, row in df.iterrows():
        values = [str(row[col]) if row[col] is not None else "" for col in df.columns]
        tree_bd.insert("", tk.END, values=values)

    log_bd(f"👁️ Visualizando tabela '{tabela}' (até 200 linhas).")

ttk.Button(frame_view_top, text="Atualizar visualização", style="Primary.TButton", command=visualizar_tabela).pack(side=tk.LEFT, padx=(5, 0))

frame_tree = tk.Frame(frame_bd_right, bg=BG_PRINCIPAL)
frame_tree.pack(fill="both", expand=True, padx=4, pady=(0, 4))

tree_bd = ttk.Treeview(frame_tree)
tree_bd.pack(side=tk.LEFT, fill="both", expand=True)

scroll_y = ttk.Scrollbar(frame_tree, orient="vertical", command=tree_bd.yview)
scroll_y.pack(side=tk.RIGHT, fill="y")
tree_bd.configure(yscrollcommand=scroll_y.set)

frame_log_bd = ttk.Labelframe(frame_bd_right, text="Log do Banco de Dados", style="Frame.TLabelframe", padding=10)
frame_log_bd.pack(fill="x", padx=4, pady=(4, 4))

txt_log_bd = tk.Text(frame_log_bd, width=60, height=8, font=("Consolas", 10),
                     bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                     bd=1, relief="solid", highlightthickness=1,
                     highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_bd.pack(fill="both", expand=True)


# =======================================================================
#                        ABA CONEXÃO BANCO DE DADOS
# =======================================================================

db_tipo_var = tk.StringVar(value="MySQL")
db_host_var = tk.StringVar()
db_port_var = tk.StringVar(value="3306")
db_user_var = tk.StringVar()
db_pass_var = tk.StringVar()
db_name_var = tk.StringVar()

lbl_conexao_title = tk.Label(frame_conexao, text="Conexão com Banco de Dados", bg=BG_PRINCIPAL, fg=FG_TEXTO, font=fonte_titulo)
lbl_conexao_title.pack(pady=(10, 10))

frame_conexao_main = tk.Frame(frame_conexao, bg=BG_PRINCIPAL)
frame_conexao_main.pack(fill="both", expand=True, padx=10, pady=10)

frame_conexao_left = tk.Frame(frame_conexao_main, bg=BG_PRINCIPAL)
frame_conexao_left.pack(side=tk.LEFT, fill="both", expand=True, padx=(0, 6))

frame_conexao_right = tk.Frame(frame_conexao_main, bg=BG_PRINCIPAL)
frame_conexao_right.pack(side=tk.LEFT, fill="both", expand=True, padx=(6, 0))

frame_conn = ttk.Labelframe(frame_conexao_left, text="Configurações de Conexão", style="Frame.TLabelframe", padding=10)
frame_conn.pack(padx=0, pady=6, fill="x")

tk.Label(frame_conn, text="Banco:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=0, column=0, padx=5, pady=5, sticky="w")
combo_tipo_db = ttk.Combobox(frame_conn, textvariable=db_tipo_var, state="readonly", width=20, values=["MySQL", "PostgreSQL"])
combo_tipo_db.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_conn, text="Host:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=1, column=0, padx=5, pady=5, sticky="w")
entry_host = tk.Entry(frame_conn, textvariable=db_host_var, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=30)
entry_host.grid(row=1, column=1, padx=5, pady=5)

tk.Label(frame_conn, text="Porta:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_port = tk.Entry(frame_conn, textvariable=db_port_var, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=30)
entry_port.grid(row=2, column=1, padx=5, pady=5)

tk.Label(frame_conn, text="Usuário:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=3, column=0, padx=5, pady=5, sticky="w")
entry_user = tk.Entry(frame_conn, textvariable=db_user_var, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=30)
entry_user.grid(row=3, column=1, padx=5, pady=5)

tk.Label(frame_conn, text="Senha:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=4, column=0, padx=5, pady=5, sticky="w")
entry_pass = tk.Entry(frame_conn, textvariable=db_pass_var, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=30, show="*")
entry_pass.grid(row=4, column=1, padx=5, pady=5)

tk.Label(frame_conn, text="Banco:", bg=BG_FRAME, fg=FG_TEXTO, font=fonte_label).grid(row=5, column=0, padx=5, pady=5, sticky="w")
entry_name = tk.Entry(frame_conn, textvariable=db_name_var, bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG, width=30)
entry_name.grid(row=5, column=1, padx=5, pady=5)

frame_log_conexao = ttk.Labelframe(frame_conexao_right, text="Log de Conexão", style="Frame.TLabelframe", padding=10)
frame_log_conexao.pack(fill="both", expand=True, padx=5, pady=5)

txt_log_conexao = tk.Text(frame_log_conexao, width=60, height=20, font=("Consolas", 10),
                          bg=INPUT_BG, fg=INPUT_FG, insertbackground=INPUT_FG,
                          bd=1, relief="solid", highlightthickness=1,
                          highlightbackground=BORDER_COR, highlightcolor=BORDER_COR)
txt_log_conexao.pack(fill="both", expand=True)

def conectar_bd():
    global db_engine, db_connected

    txt_log_conexao.delete("1.0", tk.END)

    tipo = db_tipo_var.get()
    host = db_host_var.get().strip()
    port = db_port_var.get().strip()
    user = db_user_var.get().strip()
    password = db_pass_var.get().strip()
    database = db_name_var.get().strip()

    if not (host and port and user and database):
        messagebox.showwarning("Aviso", "Preencha Host, Porta, Usuário e Banco.")
        return

    try:
        if tipo == "PostgreSQL":
            url = URL.create(
                drivername="postgresql+psycopg2",
                username=user,
                password=password,
                host=host,
                port=int(port),
                database=database
            )
        elif tipo == "MySQL":
            url = URL.create(
                drivername="mysql+pymysql",
                username=user,
                password=password,
                host=host,
                port=int(port),
                database=database
            )
        else:
            raise ValueError("Tipo de banco não suportado.")

        engine = create_engine(url)
        with engine.connect() as conn:
            conn.execute(sql_text("SELECT 1"))

        db_engine = engine
        db_connected = True
        try:
            db_status_label.config(text="BD: Conectado")
        except Exception:
            pass

        cfg = {"tipo": tipo, "host": host, "port": port, "user": user, "password": password, "database": database}
        with open(DB_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)

        txt_log_conexao.insert(tk.END, "✅ Conectado com sucesso!\n")
        txt_log_conexao.insert(tk.END, f"Configuração salva em {DB_CONFIG_FILE}\n")
        set_status("BD conectado e configuração salva.")
        messagebox.showinfo("OK", "Conectado e configuração salva com sucesso!")

    except Exception as e:
        db_engine = None
        db_connected = False
        try:
            db_status_label.config(text="BD: Desconectado")
        except Exception:
            pass
        txt_log_conexao.insert(tk.END, f"❌ Erro ao conectar:\n{e}")
        set_status("Falha ao conectar ao BD.")
        messagebox.showerror("Erro", f"Falha ao conectar.\n\n{e}")
def auto_conectar_bd():
    global db_engine, db_connected

    if not os.path.exists(DB_CONFIG_FILE):
        return

    try:
        with open(DB_CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)

        db_tipo_var.set(cfg.get("tipo", "MySQL"))
        db_host_var.set(cfg.get("host", ""))
        db_port_var.set(cfg.get("port", "3306"))
        db_user_var.set(cfg.get("user", ""))
        db_pass_var.set(cfg.get("password", ""))
        db_name_var.set(cfg.get("database", ""))

        tipo = db_tipo_var.get()
        host = db_host_var.get().strip()
        port = db_port_var.get().strip()
        user = db_user_var.get().strip()
        password = db_pass_var.get().strip()
        database = db_name_var.get().strip()

        if not (host and port and user and database):
            return

        if tipo == "PostgreSQL":
            url = URL.create("postgresql+psycopg2", username=user, password=password, host=host, port=int(port), database=database)
        elif tipo == "MySQL":
            url = URL.create("mysql+pymysql", username=user, password=password, host=host, port=int(port), database=database)
        else:
            return

        engine = create_engine(url)
        with engine.connect() as conn:
            conn.execute(sql_text("SELECT 1"))

        db_engine = engine
        db_connected = True
        try:
            db_status_label.config(text="BD: Conectado (auto)")
        except Exception:
            pass
        txt_log_conexao.insert(tk.END, "✅ Conectado automaticamente usando configuração salva.\n")
        set_status("BD conectado automaticamente.")

    except Exception as e:
        db_engine = None
        db_connected = False
        try:
            db_status_label.config(text="BD: Desconectado")
        except Exception:
            pass
        txt_log_conexao.insert(tk.END, f"⚠️ Não foi possível conectar automaticamente:\n{e}\n")
        set_status("BD não conectado (auto).")

ttk.Button(frame_conn, text="Conectar & Salvar", style="Accent.TButton", command=conectar_bd).grid(row=6, column=0, columnspan=2, pady=12)


# =======================================================================
#                           MAINLOOP
# =======================================================================

janela.after(500, auto_conectar_bd)
set_theme(current_theme_name.get())
janela.mainloop()
